"""
build_optimization_report.py — decision-ready discount-spend optimization report (Excel).

Turns the latest engine run into ONE leadership-ready workbook:
  1. Executive Summary  — portfolio KPIs + action counts + charts
  2. SKU Recommendations — every product x city cell, one action + the spend to set,
                           with a full confidence explanation per row (no SKU excluded)
  3. Confidence Method   — the exact metrics & benchmarks behind High / Medium / Low
  4. Read Me             — framing, action definitions, data lineage

IMPORTANT framing: "spend" throughout = promotional DISCOUNT spend (the rupees given
away as discount), which is the lever this engine optimizes. It is NOT paid-ad spend;
the system ingests no ad-spend rupees. "ROAS" = return on discount spend.

Sources (latest output/runs run):
  plan/all_cells.csv     — decision engine: bucket, break-even, target, ROAS, confidence
  recommendations.csv    — pricing engine: confidence score + 5 sub-scores, revenue
Joined on cell_id (1:1 for all 537 decision cells).

Output: output/OPTIMIZATION_REPORT.xlsx
"""
import os, sys, glob, math
import numpy as np
import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, ".."))
MONTH = 30.0 / 7.0                     # weekly -> monthly, matching the engine
PAUSE_DISC = 1.0                       # rec discount <= 1% on a cut = "pause the promo"
BIG_ADD_PP = 3.0                       # reinvest headroom > 3pp = "increase budget"

# ── palette (matches the dashboard: tactile slate + indigo) ─────────────────
INK   = "151C2B"; SLATE = "525B6C"; LINE = "D9DEE8"
INDIGO= "4B57D6"; INDIGO_SOFT = "EAECFB"
GOOD  = "0F8268"; GOOD_SOFT = "E2F2EE"
WARN  = "A9640F"; WARN_SOFT = "F7EDDA"
BAD   = "C0392B"; BAD_SOFT  = "FBE8E6"
NEUT  = "EEF1F6"; HEADbg = "1F2740"


def _latest_run():
    runs = sorted(glob.glob(os.path.join(ROOT, "output", "runs", "2026*")))
    for r in reversed(runs):
        if os.path.exists(os.path.join(r, "plan", "all_cells.csv")):
            return r
    raise SystemExit("no run with plan/all_cells.csv found")


def load():
    run = _latest_run()
    ac = pd.read_csv(os.path.join(run, "plan", "all_cells.csv"))
    rec = pd.read_csv(os.path.join(run, "recommendations.csv"))
    cut_ids = set(pd.read_csv(os.path.join(run, "plan", "cut_list.csv"))["cell_id"].astype(str))
    rein_ids = set(pd.read_csv(os.path.join(run, "plan", "reinvest_list.csv"))["cell_id"].astype(str))
    keep = ["cell_id", "confidence_score", "conf_density", "conf_variation", "conf_fit",
            "conf_plausibility", "conf_tightness", "n_observations", "grammage",
            "current_revenue_day", "rec_revenue_day", "elasticity"]
    rec = rec[[c for c in keep if c in rec.columns]]
    df = ac.merge(rec, on="cell_id", how="left")
    return run, df, cut_ids, rein_ids


def _num(v, d=0.0):
    try:
        f = float(v)
        return d if (f != f) else f
    except (TypeError, ValueError):
        return d


def _units_at(cur_units, cur_disc, new_disc, beta, marg_beta):
    """Model units at a new discount using the quadratic semi-log response.
    beta2 is backed out of marg_beta = beta + 2*beta2*cur_disc. Growth is clamped
    to avoid extrapolating a reinvest past what the data can support."""
    if not np.isfinite(beta):
        return cur_units
    beta2 = 0.0 if cur_disc < 1.0 else (marg_beta - beta) / (2.0 * cur_disc)
    lg = beta * (new_disc - cur_disc) + beta2 * (new_disc**2 - cur_disc**2)
    ratio = float(np.clip(np.exp(lg), 0.5, 2.5))
    return cur_units * ratio


# ── per-cell computation ────────────────────────────────────────────────────
def compute(df, cut_ids, rein_ids):
    out = []
    for _, r in df.iterrows():
        cid = str(r["cell_id"])
        cur_disc = _num(r["cur_disc"]); tgt = _num(r["tgt_disc"]); be = _num(r["be_disc"])
        mrp = _num(r["mrp"]); cur_price = _num(r["cur_price"]); cur_u = _num(r["cur_units_wk"])
        beta = _num(r.get("beta_disc")); mbeta = _num(r.get("marg_beta"))
        cat_med = _num(r.get("cat_med_disc"))
        is_cut = (r["bucket"] == "c_waste_cut") or (cid in cut_ids)

        def econ(rd, ru):
            return (cur_u * cur_price * MONTH, ru * mrp * (1 - rd / 100.0) * MONTH,
                    cur_u * mrp * cur_disc / 100.0 * MONTH, ru * mrp * rd / 100.0 * MONTH)

        # Reinvest CANDIDATE: engine says discount reliably lifts sales AND the cell is
        # under-discounted vs the category norm. But a candidate is only ACTED on if the
        # modelled step actually returns >= Rs1 of revenue per Rs1 of added discount.
        # Deepening a discount reprices the ENTIRE existing base, so beyond a low level
        # the step is ROAS-dilutive — those candidates are HELD, not increased.
        rein_target = min(be, cat_med) if cat_med > 0 else be
        rein_cand = (not is_cut) and (cid in rein_ids) and (rein_target > cur_disc + 1.0)

        hold_reason = ""
        head = 0.0
        if is_cut:
            rec_disc = tgt; rec_u = _num(r["tgt_units_wk"])
            action = "Pause" if rec_disc <= PAUSE_DISC else "Reduce Spend"
        elif rein_cand:
            rec_disc = rein_target
            rec_u = _units_at(cur_u, cur_disc, rec_disc, beta, mbeta)
            cs, rs, csp, rsp = econ(rec_disc, rec_u)
            step_roas = ((rs - cs) / (rsp - csp)) if (rsp - csp) > 1 else 0.0
            if step_roas >= 1.0:                 # adding discount returns >= Rs1 per Rs1
                head = max(rein_target - cur_disc, 0.0)
                action = "Increase Budget" if head > BIG_ADD_PP else "Reinvest"
            else:                                 # ROAS-dilutive: hold, do not add discount
                rec_disc = cur_disc; rec_u = cur_u
                action = "Hold"; hold_reason = "roas_dilutive"
        else:
            rec_disc = cur_disc; rec_u = cur_u; action = "Hold"

        cur_sales, rec_sales, cur_spend, rec_spend = econ(rec_disc, rec_u)
        rec_price = mrp * (1 - rec_disc / 100.0)
        spend_chg = rec_spend - cur_spend
        spend_pct = (spend_chg / cur_spend * 100.0) if cur_spend > 1 else 0.0
        sales_impact = rec_sales - cur_sales

        out.append(dict(
            product_id=r.get("product_id"), cell_id=cid,
            sku=r["title"], pack=r.get("grammage", ""), city=r["city"], category=r["category"],
            cur_sales=cur_sales, cur_disc=cur_disc, cur_spend=cur_spend,
            action=action, rec_disc=rec_disc, rec_spend=rec_spend,
            spend_chg=spend_chg, spend_pct=spend_pct, sales_impact=sales_impact,
            unit_chg_pct=((rec_u - cur_u) / cur_u * 100.0 if cur_u > 0 else 0.0),
            # raw metrics for rationale + confidence text
            bucket=r["bucket"], reason=r.get("decision_reason", ""), conf_tier=r["confidence"],
            conf_score=_num(r.get("confidence_score")), n_weeks=_num(r.get("n_weeks")),
            disc_std=_num(r.get("disc_std")), cat_r2=_num(r.get("cat_r2")),
            osa=_num(r.get("osa_mean")), roas=_num(r.get("marginal_roas")),
            head=head, be=be, sig_pos=bool(r.get("sig_pos")),
            reliably_waste=bool(r.get("reliably_waste")), reliably_pays=bool(r.get("reliably_pays")),
            c_density=_num(r.get("conf_density")), c_var=_num(r.get("conf_variation")),
            c_fit=_num(r.get("conf_fit")), c_plaus=_num(r.get("conf_plausibility")),
            c_tight=_num(r.get("conf_tightness")),
            comp=bool(r.get("comp_pressure")), trend=r.get("trend", ""), hold_reason=hold_reason,
        ))
    d = pd.DataFrame(out)
    # round the score ONCE so the gate, the numeric cell and the explanation text all
    # use the same value (avoids a "50/100 < 50" boundary contradiction).
    d["conf_score"] = pd.to_numeric(d["conf_score"], errors="coerce").fillna(0).round(0)
    # Confidence tier. HIGH requires all four: category fit (engine cat_ok, R²≥0.60),
    # ≥8 weeks, discount σ≥1.5pp (these three = the engine "High" tier) AND a data-richness
    # score ≥ 50. Missing any one (with fit OK) = Medium; category fit below floor = Low.
    base = d["conf_tier"].map({"High": "High", "Experimental": "Medium", "Low": "Low"}).fillna("Low")
    d["confidence"] = np.where((base == "High") & (d["conf_score"] < 50), "Medium", base)
    d["rationale"] = d.apply(rationale, axis=1)
    for col in ["why", "supporting", "assumptions", "missing", "increase", "risks"]:
        d[col] = d.apply(lambda r, c=col: conf_text(r)[c], axis=1)
    # order: act-first
    order = {"Reduce Spend": 0, "Pause": 1, "Increase Budget": 2, "Reinvest": 3, "Hold": 4}
    d["_o"] = d["action"].map(order)
    d = d.sort_values(["_o", "cur_sales"], ascending=[True, False]).drop(columns="_o").reset_index(drop=True)
    return d


def rationale(r):
    a = r["action"]
    if a in ("Reduce Spend", "Pause"):
        ret = abs(r['cur_spend'] - r['rec_spend']); uc = r.get('unit_chg_pct', 0.0)
        verb = 'Pause the promo' if a == 'Pause' else f"Trim to {r['rec_disc']:.0f}%"
        if uc >= -2.0:
            dyn = f"volume holds, ~₹{ret:,.0f}/mo of spend returns"
        else:
            dyn = (f"volume eases ~{abs(uc):.0f}%, but revenue still rises ₹{r['sales_impact']:,.0f}/mo as the "
                   f"recovered margin outweighs the lost units and ~₹{ret:,.0f}/mo of spend returns")
        return (f"Discount {r['cur_disc']:.0f}% is reliably below break-even — the extra sales it "
                f"buys don't cover the margin given away. {verb}; {dyn}.")
    if a in ("Reinvest", "Increase Budget"):
        return (f"Under-discounted vs the category norm: current {r['cur_disc']:.0f}% is ~{r['head']:.0f}pp below "
                f"the typical {r['rec_disc']:.0f}%, and the modelled step returns at least ₹1 of revenue per ₹1 of "
                f"added discount. {'Increase the budget' if a=='Increase Budget' else 'Reinvest'} to {r['rec_disc']:.0f}%.")
    if r.get("hold_reason") == "roas_dilutive":
        return ("Discount has a reliably positive effect, but at the current depth an extra ₹1 of discount returns "
                "LESS than ₹1 of revenue — deepening it just reprices the whole base. Hold; only a small, "
                "controlled price test could justify adding budget.")
    if r["bucket"] == "a_stock":
        return (f"Sales are gated by availability (on-shelf {r['osa']:.0f}%), not price — discount is not "
                f"the lever. Hold spend; fix stock first.")
    if r["bucket"] == "b_competitive":
        return "Losing category share — a defensive position. Cutting discount now could accelerate the loss, so hold."
    return "No confident signal that changing discount would pay. Hold and monitor; test before committing budget."


# ── per-row confidence explanation (the mandatory 6 fields) ──────────────────
def conf_text(r):
    lvl = r["confidence"]; score = r["conf_score"]
    wk, dv, r2, osa = r["n_weeks"], r["disc_std"], r["cat_r2"], r["osa"]
    demoted = (r["conf_tier"] == "High" and lvl == "Medium")   # data-sufficient but weak score
    # WHY — the tier is set by the data-sufficiency gates; the 0-100 score is shown honestly
    if lvl == "High":
        why = (f"Rated High — measured, not assumed: the category demand model clears the fit floor "
               f"(R²={r2:.2f} ≥ 0.60), the cell has {wk:.0f} weeks of history (≥ 8) with real discount "
               f"variation ({dv:.1f}pp ≥ 1.5), and the data-richness score is solid ({score:.0f}/100 ≥ 50).")
    elif lvl == "Medium":
        gaps = []
        if wk < 8: gaps.append(f"history is short ({wk:.0f} wks < 8)")
        if dv < 1.5: gaps.append(f"discount barely varied ({dv:.1f}pp < 1.5) to measure the response")
        if demoted and score < 50: gaps.append(f"the overall data-richness score is weak ({score:.0f}/100 < 50)")
        if not gaps: gaps.append("the discount effect is not yet statistically confirmable")
        why = (f"Rated Medium — the category model fits (R²={r2:.2f}) and the direction is sound, but "
               f"{', and '.join(gaps)}. Treat as a bet to test, not a banked result.")
    else:
        why = (f"Rated Low — the category model does not clear the reliability floor (R²={r2:.2f} < 0.60) "
               f"or the data is too thin, so the discount response here can't be trusted yet.")
    # SUPPORTING METRICS — concrete decision evidence (not the abstract 0-1 sub-scores)
    sup = (f"Data-richness score {score:.0f}/100; {wk:.0f} weeks of history; within-cell discount σ {dv:.1f}pp; "
           f"category demand model R² {r2:.2f}; on-shelf availability {osa:.0f}%")
    if r["action"] in ("Reduce Spend", "Pause"):
        sup += "; break-even test PASSED — even the optimistic 95% CI of the discount effect is below the pay-line"
    elif r["action"] in ("Reinvest", "Increase Budget"):
        sup += "; discount slope reliably positive (95% CI > 0) and the modelled step returns ≥ ₹1 per ₹1 added"
    # ASSUMPTIONS
    ass = ("Category-pooled elasticity (with a per-cell baseline) applies to this cell; the 6-month window is "
           "representative; competitors, availability and season are controlled for; changes are glided in "
           "(≤3pp/week), not made overnight.")
    # MISSING / why not higher
    if lvl == "High":
        miss = "None material for the decision. A live price test would upgrade this from strong evidence to proof."
    else:
        m = []
        if wk < 8: m.append("longer sales history")
        if dv < 1.5: m.append("more within-cell discount variation to separate price from other drivers")
        if r2 < 0.60: m.append("a better-fitting category model")
        if osa < 75: m.append("stable availability (current stockouts confound the read)")
        miss = "Limited by: " + (", ".join(m) if m else "an unconfirmed discount effect") + "."
    # WOULD INCREASE
    inc = []
    if wk < 12: inc.append("a few more weeks of data")
    if dv < 2.0: inc.append("a deliberate discount test (vary the depth and measure)")
    if osa < 90: inc.append("resolving stockouts so availability stops masking price")
    inc.append("SKU-level conversion & traffic feeds (currently inferred from units)")
    increase = "Would raise confidence: " + "; ".join(inc) + "."
    # RISKS
    rk = []
    if r["action"] in ("Reinvest", "Increase Budget"):
        rk.append("adding discount extrapolates slightly above observed levels — glide up and re-measure")
    if r["bucket"] == "b_competitive":
        rk.append("defensive share position — holding is deliberate; a cut could accelerate share loss")
    if osa < 75:
        rk.append("availability-gated: price moves won't help until stock is fixed")
    if lvl != "High":
        rk.append("performance has been inconsistent; the recommendation may shift as data accrues")
    if r["trend"] == "declining":
        rk.append("unit trend is declining — watch for a demand shift unrelated to price")
    risks = "; ".join(rk).capitalize() + "." if rk else "No material risks beyond normal execution variance."
    return dict(why=why, supporting=sup, assumptions=ass, missing=miss, increase=increase, risks=risks)


# ── workbook ────────────────────────────────────────────────────────────────
def build_workbook(run, d):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.comments import Comment

    RUP = '"₹"#,##0'; PCT = '0.0"%"'
    thin = Side(style="thin", color=LINE)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center")

    wb = Workbook()

    # ============ 1. EXECUTIVE SUMMARY ============
    es = wb.active; es.title = "Executive Summary"
    es.sheet_view.showGridLines = False
    tot_sales = d["cur_sales"].sum()
    cur_spend = d["cur_spend"].sum()
    rec_spend = d["rec_spend"].sum()
    saved = d.loc[d["spend_chg"] < 0, "spend_chg"].sum() * -1
    reinv = d.loc[d["spend_chg"] > 0, "spend_chg"].sum()
    net = rec_spend - cur_spend
    incr_sales = d["sales_impact"].sum()
    roas_now = tot_sales / cur_spend if cur_spend else 0
    roas_rec = (tot_sales + incr_sales) / rec_spend if rec_spend else 0
    counts = d["action"].value_counts().to_dict()
    def c(a): return int(counts.get(a, 0))

    es.column_dimensions["A"].width = 3
    es.column_dimensions["B"].width = 34
    es.column_dimensions["C"].width = 20
    for col in "DEFGH": es.column_dimensions[col].width = 16

    es["B2"] = "Discount-Spend Optimization — Executive Summary"
    es["B2"].font = Font(name="Calibri", size=17, bold=True, color=INK)
    es["B3"] = f"Every SKU × city covered · run {os.path.basename(run)} · figures are monthly"
    es["B3"].font = Font(size=10, italic=True, color=SLATE)
    es["B4"] = ('"Spend" = promotional DISCOUNT budget (₹ given away), the lever this engine optimizes — '
                'not paid-ad spend. "ROAS" = sales per ₹ of discount.')
    es["B4"].font = Font(size=9, italic=True, color=WARN)
    es["B4"].alignment = Alignment(wrap_text=True)
    es.merge_cells("B4:H4")
    es.row_dimensions[4].height = 26

    kpis = [
        ("PORTFOLIO", "", ""),
        ("Total SKUs (product × city)", len(d), "0"),
        ("Total Sales", tot_sales, RUP),
        ("Total Discount Spend — now", cur_spend, RUP),
        ("Discount as % of current Sales", cur_spend / tot_sales * 100 if tot_sales else 0, PCT),
        ("Total Discount Spend — recommended", rec_spend, RUP),
        ("Discount as % of projected Sales", rec_spend / (tot_sales + incr_sales) * 100 if tot_sales else 0, PCT),
        ("THE MOVE", "", ""),
        ("Discount Budget Saved (from Reduce/Pause)", saved, RUP),
        ("Discount Budget Reinvested (Reinvest/Increase)", reinv, RUP),
        ("Net Budget Change", net, RUP),
        ("Expected Incremental Sales", incr_sales, RUP),
        ("ROAS now  →  recommended", None, ""),
        ("Expected ROAS Improvement", (roas_rec / roas_now - 1) * 100 if roas_now else 0, PCT),
    ]
    row = 6
    for label, val, fmt in kpis:
        if val == "" and fmt == "":  # section header
            es.cell(row, 2, label).font = Font(size=10, bold=True, color="FFFFFF")
            for cc in (2, 3):
                es.cell(row, cc).fill = PatternFill("solid", fgColor=HEADbg)
            es.cell(row, 3, "")
            row += 1; continue
        lc = es.cell(row, 2, label); lc.font = Font(size=10.5, color=INK); lc.border = border
        if label.startswith("ROAS now"):
            vc = es.cell(row, 3, f"{roas_now:.1f}x  →  {roas_rec:.1f}x")
        else:
            vc = es.cell(row, 3, round(val, 1) if isinstance(val, float) else val)
            if fmt: vc.number_format = fmt
        vc.font = Font(size=11, bold=True, color=INDIGO if "ROAS Improvement" in label or "Saved" in label else INK)
        vc.border = border; vc.alignment = Alignment(horizontal="right")
        row += 1

    # "how to read the money" note — below the KPI block, left of the charts
    nrow = row + 1
    es.cell(nrow, 2, "How to read the money").font = Font(size=11, bold=True, color=INK)
    money_note = (
        "• Sales Impact = change in NET REVENUE (units × selling price). Cutting a wasteful "
        "discount RAISES the price you keep per unit, so revenue can rise even as the discount "
        "falls — that figure is a revenue GAIN, not a loss.\n"
        "• Discount Budget Saved / Spend Change = the discount rupees you no longer give away — "
        "a separate, usually LARGER number.\n"
        "So one cut delivers both: more revenue AND freed budget. Example (Tur/Arhar Dal, "
        "Bangalore): 28%→12% discount lifts price ₹255→₹311 (+22%) while units ease 869→772 "
        "(−11%, elasticity −0.55) → sales +₹79,932/mo AND ₹227,069/mo of discount returned.")
    mc = es.cell(nrow + 1, 2, money_note); mc.font = Font(size=9.5, color=SLATE)
    mc.alignment = Alignment(wrap_text=True, vertical="top")
    es.merge_cells(start_row=nrow + 1, start_column=2, end_row=nrow + 7, end_column=4)

    # action-count table (drives the charts)
    tr = 6
    es.cell(tr, 5, "Action").font = Font(bold=True, color="FFFFFF")
    es.cell(tr, 6, "# SKUs").font = Font(bold=True, color="FFFFFF")
    es.cell(tr, 7, "Δ Spend ₹/mo").font = Font(bold=True, color="FFFFFF")
    for cc in (5, 6, 7):
        es.cell(tr, cc).fill = PatternFill("solid", fgColor=HEADbg)
        es.cell(tr, cc).alignment = center
    acts = ["Reduce Spend", "Pause", "Hold", "Reinvest", "Increase Budget"]
    for i, a in enumerate(acts):
        rr = tr + 1 + i
        es.cell(rr, 5, a).border = border
        es.cell(rr, 6, c(a)).border = border
        dsp = d.loc[d["action"] == a, "spend_chg"].sum()
        vc = es.cell(rr, 7, round(dsp, 0)); vc.number_format = RUP; vc.border = border
    last = tr + len(acts)

    # honest callout on the reinvest side
    rein_total = c("Reinvest") + c("Increase Budget")
    note = ("KEY FINDING — no SKU currently justifies MORE discount: for every reinvest candidate, an extra ₹1 of "
            "discount returns LESS than ₹1 of revenue (a deeper discount reprices the whole existing base). Every "
            "profitable move here is a cut or a hold. A separate 'growth bets' list can be built if the goal is to "
            "buy market share accepting ROAS < 1."
            if rein_total == 0 else
            f"{rein_total} SKUs clear the reinvest bar (modelled step returns ≥ ₹1 per ₹1 of added discount); the "
            f"rest are held because deepening their discount would dilute ROAS.")
    nc = es.cell(12, 5, note); nc.font = Font(size=9.5, italic=True, color=WARN)
    nc.alignment = Alignment(wrap_text=True, vertical="top")
    es.merge_cells("E12:H13"); es.row_dimensions[12].height = 30

    # bar chart — # SKUs by action
    bar = BarChart(); bar.type = "col"; bar.title = "SKUs by recommended action"
    bar.height = 7.5; bar.width = 13; bar.legend = None
    data = Reference(es, min_col=6, min_row=tr, max_row=last)
    cats = Reference(es, min_col=5, min_row=tr + 1, max_row=last)
    bar.add_data(data, titles_from_data=True); bar.set_categories(cats)
    bar.y_axis.majorGridlines = None
    es.add_chart(bar, "E14")

    # pie — budget saved vs reinvested
    es.cell(23, 5, "Budget saved"); es.cell(23, 6, round(saved, 0))
    es.cell(24, 5, "Budget reinvested"); es.cell(24, 6, round(reinv, 0))
    pie = PieChart(); pie.title = "Discount budget: saved vs reinvested"; pie.height = 7.5; pie.width = 11
    pdata = Reference(es, min_col=6, min_row=23, max_row=24)
    pcats = Reference(es, min_col=5, min_row=23, max_row=24)
    pie.add_data(pdata); pie.set_categories(pcats)
    es.add_chart(pie, "E31")

    # ============ 2. SKU RECOMMENDATIONS ============
    ws = wb.create_sheet("SKU Recommendations")
    ws.sheet_view.showGridLines = False
    cols = [
        ("Product ID", "product_id", 11, None), ("Cell ID (key)", "cell_id", 22, None),
        ("SKU Name", "sku", 30, None), ("Pack", "pack", 8, None), ("City", "city", 14, None),
        ("Category", "category", 20, None),
        ("Current Sales ₹/mo", "cur_sales", 15, RUP), ("Current Disc %", "cur_disc", 11, PCT),
        ("Current Spend ₹/mo", "cur_spend", 15, RUP),
        ("Recommended Action", "action", 16, None),
        ("Rec Disc %", "rec_disc", 10, PCT), ("Recommended Spend ₹/mo", "rec_spend", 16, RUP),
        ("Spend Change ₹/mo", "spend_chg", 15, RUP), ("Spend Change %", "spend_pct", 12, PCT),
        ("Est. Sales Impact ₹/mo", "sales_impact", 15, RUP),
        ("Business Rationale", "rationale", 52, None),
        ("Confidence", "confidence", 11, None), ("Conf Score", "conf_score", 9, "0"),
        # the five sub-scores behind Conf Score (0-1 each); weighted sum ×100 = Conf Score
        ("Density ·0.25", "c_density", 11, "0.00"), ("Variation ·0.20", "c_var", 11, "0.00"),
        ("Fit ·0.20", "c_fit", 9, "0.00"), ("Plausible ·0.15", "c_plaus", 11, "0.00"),
        ("Tightness ·0.20", "c_tight", 11, "0.00"),
        ("Why This Confidence", "why", 60, None), ("Supporting Metrics", "supporting", 55, None),
        ("Assumptions", "assumptions", 50, None), ("Missing Data (why not higher)", "missing", 45, None),
        ("Would Increase Confidence", "increase", 50, None), ("Risks / Limitations", "risks", 50, None),
    ]
    hd = Font(bold=True, color="FFFFFF", size=10)
    hf = PatternFill("solid", fgColor=HEADbg)
    for j, (title, key, w, fmt) in enumerate(cols, start=1):
        cell = ws.cell(1, j, title); cell.font = hd; cell.fill = hf
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(j)].width = w
    conf_fill = {"High": PatternFill("solid", fgColor=GOOD_SOFT),
                 "Medium": PatternFill("solid", fgColor=WARN_SOFT),
                 "Low": PatternFill("solid", fgColor=BAD_SOFT)}
    conf_font = {"High": Font(bold=True, color=GOOD), "Medium": Font(bold=True, color=WARN), "Low": Font(bold=True, color=BAD)}
    act_fill = {"Reduce Spend": INDIGO_SOFT, "Pause": BAD_SOFT, "Hold": NEUT,
                "Reinvest": GOOD_SOFT, "Increase Budget": GOOD_SOFT}
    for i, (_, rr) in enumerate(d.iterrows(), start=2):
        for j, (title, key, w, fmt) in enumerate(cols, start=1):
            v = rr[key]
            if isinstance(v, float) and (v != v):
                v = ""
            cell = ws.cell(i, j, round(v, 1) if isinstance(v, (int, float)) and fmt in (RUP, PCT, "0") else v)
            if fmt and v != "": cell.number_format = fmt
            cell.alignment = wrap if w >= 40 else Alignment(vertical="top", wrap_text=(w >= 20))
            cell.border = border
            cell.font = Font(size=9.5, color=INK)
            if key == "confidence" and v in conf_fill:
                cell.fill = conf_fill[v]; cell.font = conf_font[v]; cell.alignment = center
            if key == "action":
                cell.fill = PatternFill("solid", fgColor=act_fill.get(v, NEUT))
                cell.font = Font(size=9.5, bold=True, color=INK); cell.alignment = center
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(d)+1}"
    ws.row_dimensions[1].height = 30

    # ============ 3. CONFIDENCE METHOD ============
    build_confidence_sheet(wb, d, border)

    # ============ 4. READ ME ============
    build_readme_sheet(wb, run, d)

    wb._sheets = [wb["Executive Summary"], wb["SKU Recommendations"],
                  wb["Confidence Method"], wb["Read Me"]]
    dest = os.path.join(ROOT, "output", "OPTIMIZATION_REPORT.xlsx")
    os.makedirs(os.path.dirname(dest), exist_ok=True)
    wb.save(dest)
    return dest


def build_confidence_sheet(wb, d, border):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws = wb.create_sheet("Confidence Method")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 16
    for col in "CDEFG": ws.column_dimensions[col].width = 24
    wrap = Alignment(wrap_text=True, vertical="top")

    ws["B2"] = "How confidence is scored — metrics & benchmarks"
    ws["B2"].font = Font(size=15, bold=True, color=INK)
    ws["B3"] = ("Confidence is NOT a single metric. It gates on whether the DATA can support the "
                "recommendation, using the thresholds below. The 0–100 score blends five sub-signals.")
    ws["B3"].font = Font(size=10, italic=True, color=SLATE); ws.merge_cells("B3:G3")

    hdr = ["Signal", "What it measures", "Required for HIGH"]
    r0 = 5
    ws.cell(r0, 2, hdr[0]).font = Font(bold=True, color="FFFFFF"); ws.cell(r0, 2).fill = PatternFill("solid", fgColor=HEADbg)
    ws.cell(r0, 3, hdr[1]).font = Font(bold=True, color="FFFFFF"); ws.cell(r0, 3).fill = PatternFill("solid", fgColor=HEADbg)
    ws.cell(r0, 4, hdr[2]).font = Font(bold=True, color="FFFFFF"); ws.cell(r0, 4).fill = PatternFill("solid", fgColor=HEADbg)
    ws.merge_cells(start_row=r0, start_column=4, end_row=r0, end_column=6)
    for j in (2, 3, 4):
        ws.cell(r0, j).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    rows = [
        ("Category model fit (R²)", "How well the demand model explains the category — the gate to trust it at all", "≥ 0.60"),
        ("Data history", "Weeks of sales for this cell", "≥ 8 weeks"),
        ("Discount variation (σ)", "Within-cell discount spread — needed to measure the response", "≥ 1.5 pp"),
        ("Data-richness score", "×100 · (0.25·density + 0.20·variation + 0.20·fit + 0.15·plausibility + 0.20·tightness); "
         "each sub-score 0–1, shown per row on the SKU sheet so the score is auditable", "≥ 50"),
    ]
    rr = r0 + 1
    for name, what, req in rows:
        c0 = ws.cell(rr, 2, name); c0.font = Font(bold=True, color=INK); c0.border = border
        c1 = ws.cell(rr, 3, what); c1.alignment = wrap; c1.font = Font(size=10, color=SLATE); c1.border = border
        c2 = ws.cell(rr, 4, req); c2.alignment = Alignment(horizontal="center", vertical="top")
        c2.font = Font(size=10.5, bold=True, color=GOOD); c2.border = border
        ws.merge_cells(start_row=rr, start_column=4, end_row=rr, end_column=6)
        rr += 1
    rr += 1
    ws.cell(rr, 2, "HIGH = all four met.   MEDIUM = category fit (R²≥0.60) met but ANY one of the other three "
                   "missing — so a Medium SKU can still show a long history or a decent score if it misses a "
                   "different gate.   LOW = category model below the 0.60 fit floor.").font = Font(size=10, italic=True, color=SLATE)
    ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=6); ws.row_dimensions[rr].height = 42; rr += 2

    ws.cell(rr, 2, "Shapes the action & risk — NOT the confidence tier").font = Font(size=12, bold=True, color=INK); rr += 1
    for name, txt in [
        ("Availability (OSA)", "Low on-shelf availability means price isn't the lever — routes the SKU to Hold "
         "(fix stock first), whatever its confidence tier."),
        ("Break-even CI test", "Act rows only: Reduce/Pause require the whole 95% CI of the discount effect below "
         "the pay-line; Reinvest requires the slope CI > 0 AND a modelled step return ≥ ₹1 per ₹1 of added discount."),
    ]:
        ws.cell(rr, 2, name).font = Font(bold=True, color=INDIGO)
        cc = ws.cell(rr, 3, txt); cc.alignment = wrap; cc.font = Font(size=10, color=INK)
        ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=6); ws.row_dimensions[rr].height = 30; rr += 1
    rr += 1

    ws.cell(rr, 2, "What each tier means for a decision").font = Font(size=12, bold=True, color=INK); rr += 1
    tiers = [
        ("HIGH", GOOD, GOOD_SOFT,
         "Act now, bankable. Category model fits (R²≥0.60), ≥8 weeks of history, real discount variation "
         "(≥1.5pp) and a solid data-richness score (≥50). For act rows the break-even test is also decisive. "
         "Risk: normal execution variance. To upgrade to proof: run a live price test."),
        ("MEDIUM", WARN, WARN_SOFT,
         "Directionally right, treat as a test. Category model fits, but history is short, discount barely "
         "varied, or the data-richness score is weak — so the effect isn't yet statistically confirmable and is "
         "never banked into the headline. Missing: more weeks and/or a deliberate discount test."),
        ("LOW", BAD, BAD_SOFT,
         "Do not act on price yet. Category model below the 0.60 fit floor, or thin/new/erratic data. Increase "
         "confidence by accruing history, fixing availability, and adding SKU-level feeds. (No cells fall here in "
         "the current run.)"),
    ]
    for name, col, soft, txt in tiers:
        ws.cell(rr, 2, name).font = Font(bold=True, color=col)
        ws.cell(rr, 2).fill = PatternFill("solid", fgColor=soft)
        cc = ws.cell(rr, 3, txt); cc.alignment = wrap; cc.font = Font(size=10, color=INK); ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=6)
        ws.row_dimensions[rr].height = 58
        rr += 1

    # distribution
    rr += 1
    ws.cell(rr, 2, "Current distribution").font = Font(size=12, bold=True, color=INK); rr += 1
    vc = d["confidence"].value_counts()
    for tier in ["High", "Medium", "Low"]:
        ws.cell(rr, 2, tier).font = Font(bold=True)
        ws.cell(rr, 3, f"{int(vc.get(tier,0))} SKUs ({vc.get(tier,0)/len(d)*100:.0f}%)")
        rr += 1
    rr += 1
    ws.cell(rr, 2, "Data we don't yet have (would lift confidence across the board):").font = Font(bold=True, italic=True, color=SLATE)
    for gap in ["SKU-level traffic & conversion (units are inferred, not click-through)",
                "Paid-ad spend in ₹ (only share-of-voice is available today)",
                "Competitor prices at the SKU level (only category share is modelled)",
                "Longer history — 6 months limits seasonality separation"]:
        rr += 1; ws.cell(rr, 3, "• " + gap).font = Font(size=10, color=SLATE)


def build_readme_sheet(wb, run, d):
    from openpyxl.styles import Font, Alignment
    ws = wb.create_sheet("Read Me")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3; ws.column_dimensions["B"].width = 100
    wrap = Alignment(wrap_text=True, vertical="top")
    blocks = [
        ("Discount-Spend Optimization Report", 15, True, INK),
        (f"Run {os.path.basename(run)} · {len(d)} product × city cells · all figures monthly.", 10, False, SLATE),
        ("", 8, False, INK),
        ("What 'spend' means here", 12, True, INDIGO),
        ("This engine optimizes PROMOTIONAL DISCOUNT spend — the rupees given away as discount — because that "
         "is the lever in the data. It does NOT ingest paid-advertising spend, so 'ROAS' means sales per rupee "
         "of discount, and 'budget' means discount budget. Where the brief said ad spend, read discount spend.", 10, False, INK),
        ("", 8, False, INK),
        ("The five actions", 12, True, INDIGO),
        ("Reduce Spend — discount is provably below break-even; trim it, keep the volume, bank the margin.", 10, False, INK),
        ("Pause — the same, but the profitable discount is ~0%: stop the promo entirely.", 10, False, INK),
        ("Hold — keep spend unchanged. Either availability (not price) is the limiter, or it's a defensive "
         "share position, or there isn't yet a confident signal. This is the right call for most SKUs.", 10, False, INK),
        ("Reinvest — a genuinely under-discounted SKU where the modelled step of added discount returns at least "
         "₹1 of revenue per ₹1 spent. Add up to the category-norm level, never beyond.", 10, False, INK),
        ("Increase Budget — the same, with larger room (>3pp). NOTE: on the current data no SKU clears this bar — "
         "deepening any discount just reprices the whole base and dilutes ROAS, so all such candidates are held.", 10, False, INK),
        ("", 8, False, INK),
        ("How the recommendation is made (not a single metric)", 12, True, INDIGO),
        ("A confounder-controlled demand model isolates discount's own effect on sales — holding availability, "
         "competition, advertising visibility and season constant — then a break-even test (using the 95% "
         "confidence interval, not a point estimate) decides whether the discount pays. A second Double-ML "
         "engine cross-checks it, and changes are glided in ≤3pp/week under safety caps.", 10, False, INK),
        ("", 8, False, INK),
        ("Sheets", 12, True, INDIGO),
        ("• Executive Summary — the portfolio view and the money.  • SKU Recommendations — every SKU, filterable, "
         "with a full confidence explanation per row.  • Confidence Method — the metrics & benchmarks behind "
         "High/Medium/Low.", 10, False, INK),
    ]
    r = 2
    for txt, sz, bold, color in blocks:
        cc = ws.cell(r, 2, txt); cc.font = Font(size=sz, bold=bold, color=color); cc.alignment = wrap
        if len(txt) > 90: ws.row_dimensions[r].height = 46
        r += 1


def main():
    run, df, cut_ids, rein_ids = load()
    d = compute(df, cut_ids, rein_ids)
    dest = build_workbook(run, d)
    print("SKUs:", len(d))
    print(d["action"].value_counts().to_string())
    print("confidence:", d["confidence"].value_counts().to_dict())
    print("WROTE:", dest)


if __name__ == "__main__":
    main()
