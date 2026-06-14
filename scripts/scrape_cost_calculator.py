"""
scrape_cost_calculator.py — Generate a live, formula-driven Excel calculator
for the cross-platform price-tracking scraping economics.

Edit the yellow INPUT cells in the workbook and every output (scrapes/day,
monthly volume, infra cost, gross margin) recomputes via real Excel formulas
— no need to re-run this script. Re-run only if you want a fresh blank copy.

Output: SCRAPE_COST_CALCULATOR.xlsx at the project root.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, ".."))
OUT = os.path.join(ROOT, "SCRAPE_COST_CALCULATOR.xlsx")

# ── styles ────────────────────────────────────────────────────────────
TITLE_FONT   = Font(bold=True, size=14, color="FFFFFF")
HDR_FONT     = Font(bold=True, size=11, color="FFFFFF")
LBL_FONT     = Font(size=10)
VAL_FONT     = Font(size=10, bold=True)
CALC_FONT    = Font(size=10, bold=True, color="1F4E78")
NOTE_FONT    = Font(size=9, italic=True, color="808080")

TITLE_FILL   = PatternFill("solid", fgColor="1F4E78")
HDR_FILL     = PatternFill("solid", fgColor="2E75B6")
INPUT_FILL   = PatternFill("solid", fgColor="FFF2CC")   # yellow = editable
CALC_FILL    = PatternFill("solid", fgColor="E2EFDA")   # green  = computed
TIER_FILL    = PatternFill("solid", fgColor="FCE4D6")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

RUPEE = u'₹'  # ₹
CUR = '#,##0'
CUR0 = f'"{RUPEE}"#,##0'
PCT = '0.0%'
NUM = '#,##0'


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Calculator"
    ws.sheet_view.showGridLines = False

    # column widths
    widths = {"A": 2, "B": 42, "C": 16, "D": 16, "E": 16, "F": 16, "G": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    def put(cell, value, font=None, fill=None, fmt=None, align=None, border=False):
        c = ws[cell]
        c.value = value
        if font: c.font = font
        if fill: c.fill = fill
        if fmt: c.number_format = fmt
        if align: c.alignment = Alignment(horizontal=align, vertical="center")
        if border: c.border = BORDER
        return c

    # ── Title ───────────────────────────────────────────────────────
    ws.merge_cells("B1:G1")
    put("B1", "Cross-Platform Price-Tracking — Scrape Cost & Margin Calculator",
        TITLE_FONT, TITLE_FILL, align="left")
    ws.row_dimensions[1].height = 26
    put("B2", "Edit the yellow cells. Everything else updates automatically.",
        NOTE_FONT)

    # ── INPUTS ──────────────────────────────────────────────────────
    put("B4", "INPUTS  (edit these)", HDR_FONT, HDR_FILL)
    put("C4", "", HDR_FONT, HDR_FILL)
    inputs = [
        ("B5",  "SKUs",                                      90),
        ("B6",  "Variants per SKU",                          3),
        ("B7",  "Marketplace platforms (national price)",    3),
        ("B8",  "Quick-commerce platforms (per-city price)", 3),
        ("B9",  "Cities / locations tracked (q-commerce)",   10),
        ("B10", "Hours between scrapes",                     4),
        ("B11", "Days per month",                            30),
        ("B12", f"Cost per 1,000 requests ({RUPEE})",        40),
        ("B13", f"Subscription price to brand ({RUPEE}/mo)", 100000),
    ]
    for lbl_cell, lbl, val in inputs:
        put(lbl_cell, lbl, LBL_FONT)
        vcell = f"C{lbl_cell[1:]}"
        put(vcell, val, VAL_FONT, INPUT_FILL, fmt=NUM, align="center", border=True)

    # ── RESULTS ─────────────────────────────────────────────────────
    put("B15", "RESULTS  (auto-calculated)", HDR_FONT, HDR_FILL)
    put("C15", "", HDR_FONT, HDR_FILL)
    results = [
        ("B16", "Total items tracked",                 "=C5*C6",            NUM),
        ("B17", "Sweeps per day",                       "=24/C10",           '0.0'),
        ("B18", "Checks per sweep — marketplaces",      "=C16*C7*1",         NUM),
        ("B19", "Checks per sweep — quick-commerce",    "=C16*C8*C9",        NUM),
        ("B20", "Checks per sweep — TOTAL",             "=C18+C19",          NUM),
        ("B21", "Checks per DAY",                       "=C20*C17",          NUM),
        ("B22", "Checks per MONTH",                     "=C21*C11",          NUM),
        ("B23", f"Monthly infra cost ({RUPEE})",        "=C22/1000*C12",     CUR0),
        ("B24", f"Gross margin ({RUPEE}/month)",        "=C13-C23",          CUR0),
        ("B25", "Gross margin %",                       "=IF(C13=0,0,C24/C13)", PCT),
        ("B26", f"Infra cost per item ({RUPEE}/mo)",    "=IF(C16=0,0,C23/C16)", CUR0),
    ]
    for lbl_cell, lbl, formula, fmt in results:
        put(lbl_cell, lbl, LBL_FONT)
        vcell = f"C{lbl_cell[1:]}"
        put(vcell, formula, CALC_FONT, CALC_FILL, fmt=fmt, align="center", border=True)

    # ── FREQUENCY SENSITIVITY ───────────────────────────────────────
    r = 28
    put(f"B{r}", "FREQUENCY SENSITIVITY  (at current items / platforms / cities / cost)",
        HDR_FONT, HDR_FILL)
    for col in ("C", "D", "E", "F"):
        put(f"{col}{r}", "", HDR_FONT, HDR_FILL)
    r += 1
    headers = ["Scrape every…", "Sweeps/day", "Checks/month",
               f"Infra {RUPEE}/mo", f"Margin {RUPEE}/mo", "Margin %"]
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        put(f"{col}{r}", h, Font(bold=True, size=9), PatternFill("solid", fgColor="DDEBF7"),
            align="center", border=True)
    r += 1
    freqs = [("Every 1 hour", 1), ("Every 2 hours", 2), ("Every 4 hours", 4),
             ("Every 6 hours", 6), ("Every 8 hours", 8), ("Every 12 hours", 12),
             ("Once a day", 24)]
    for lbl, h in freqs:
        put(f"B{r}", lbl, LBL_FONT, border=True)
        # sweeps/day = 24/h
        put(f"C{r}", f"=24/{h}", CALC_FONT, fmt='0.0', align="center", border=True)
        # checks/month = total-per-sweep * sweeps/day * days
        put(f"D{r}", f"=$C$20*(24/{h})*$C$11", CALC_FONT, fmt=NUM, align="center", border=True)
        # infra = checks/month /1000 * cost per 1000
        put(f"E{r}", f"=D{r}/1000*$C$12", CALC_FONT, fmt=CUR0, align="center", border=True)
        # margin = price - infra
        put(f"F{r}", f"=$C$13-E{r}", CALC_FONT, fmt=CUR0, align="center", border=True)
        # margin %
        put(f"G{r}", f"=IF($C$13=0,0,F{r}/$C$13)", CALC_FONT, fmt=PCT, align="center", border=True)
        r += 1

    # ── COST-TIER COMPARISON ────────────────────────────────────────
    r += 1
    put(f"B{r}", "COLLECTION-METHOD COST TIERS  (at current frequency & volume)",
        HDR_FONT, HDR_FILL)
    for col in ("C", "D", "E", "F"):
        put(f"{col}{r}", "", HDR_FONT, HDR_FILL)
    r += 1
    tier_hdr = ["Method", f"{RUPEE}/1,000", f"Infra {RUPEE}/mo",
                f"Margin {RUPEE}/mo", "Margin %"]
    for i, h in enumerate(tier_hdr):
        col = get_column_letter(2 + i)
        put(f"{col}{r}", h, Font(bold=True, size=9), PatternFill("solid", fgColor="DDEBF7"),
            align="center", border=True)
    r += 1
    tiers = [("Cheap — datacenter proxy + JSON API (higher ban risk)", 15),
             ("Mid — residential proxy + JSON API", 40),
             ("Premium — managed unblocker / browser", 120)]
    for lbl, rate in tiers:
        put(f"B{r}", lbl, LBL_FONT, TIER_FILL, border=True)
        put(f"C{r}", rate, VAL_FONT, TIER_FILL, fmt=NUM, align="center", border=True)
        put(f"D{r}", f"=$C$22/1000*C{r}", CALC_FONT, TIER_FILL, fmt=CUR0, align="center", border=True)
        put(f"E{r}", f"=$C$13-D{r}", CALC_FONT, TIER_FILL, fmt=CUR0, align="center", border=True)
        put(f"F{r}", f"=IF($C$13=0,0,E{r}/$C$13)", CALC_FONT, TIER_FILL, fmt=PCT, align="center", border=True)
        r += 1

    # ── notes ───────────────────────────────────────────────────────
    r += 1
    notes = [
        "Notes:",
        "• Marketplaces (Amazon/Flipkart/JioMart) price nationally → counted at 1 location.",
        "• Quick-commerce (Blinkit/Zepto/Instamart) prices are per-city → multiplied by cities.",
        "• 'Checks' = one price read for one item on one platform in one location.",
        "• Cities is the biggest cost lever — keep to your priority 5-10, not all of India.",
        "• Use JSON private APIs (cheap) wherever possible; browsers cost 30-100x more.",
        "• Flat frequency is wasteful — tiering (hero SKUs frequent, tail daily) cuts 50-70% more.",
    ]
    for n in notes:
        put(f"B{r}", n, NOTE_FONT)
        r += 1

    wb.save(OUT)
    print(f"Saved: {OUT}")


if __name__ == "__main__":
    main()
