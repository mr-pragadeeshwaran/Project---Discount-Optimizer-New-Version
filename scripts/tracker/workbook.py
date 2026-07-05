"""
workbook.py — Weekly Discount Tracker: Excel workbook builder module.

Renders the weekly discount plan + guardrails + accuracy scorecard + seasonality
+ a plain-English "How to use" sheet into a single polished .xlsx file that a
NON-TECHNICAL business owner can open, read, and act on.

Public function:
    build_workbook(tracker_df, guardrail_summary, scorecard, season_info,
                   out_path, week_label) -> out_path

Dependencies: pandas / numpy / openpyxl / python stdlib only. No network.

SHARED CONTRACT NOTE
--------------------
This module is the *presentation* layer. It consumes a tracker_df that carries
the plan_df columns described in the shared contract plus a few view-friendly
aliases produced by upstream modules. To stay robust against small naming
differences between upstream modules, every column read here goes through
_pick() / _col() helpers that accept a list of candidate names and fall back
gracefully. The canonical names it looks for first are the shared-contract ones:

    Product   <- title
    City      <- city
    Category  <- category
    Current Price     <- cur_price
    Current Disc%     <- cur_disc
    This-Week Disc%   <- suggested_disc  (aliases: week_disc, this_week_disc)
    This-Week Price   <- suggested_price (aliases: week_price, this_week_price)
    Action            <- week_action     (derived from bucket if absent)
    Confidence        <- confidence
    Reason            <- decision_reason (alias: reason)
    Weekly Saving     <- weekly_saving   (alias: net_gain_wk; else derived
                                          from pred_net_rev_delta_wk)

week_action vocabulary used for row coloring: 'cut' | 'reinvest' | 'hold'.
Bucket -> action mapping when week_action is missing:
    c_waste_cut   -> cut
    e_reinvest    -> reinvest
    a_stock / b_competitive / f_monitor -> hold
"""

from __future__ import annotations

import os
import tempfile
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------- #
# Palette / style constants
# --------------------------------------------------------------------------- #
HEADER_FILL = PatternFill("solid", fgColor="1F4E5F")   # dark teal
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

TITLE_FONT = Font(bold=True, size=15, color="1F4E5F")
SUBTITLE_FONT = Font(italic=True, size=10, color="555555")
LABEL_FONT = Font(bold=True, size=11, color="1F4E5F")
BODY_FONT = Font(size=11, color="222222")

# Row / status fills (light, print-friendly)
FILL_AMBER = PatternFill("solid", fgColor="FCE8B2")    # cut
FILL_BLUE = PatternFill("solid", fgColor="D6E4F0")     # reinvest
FILL_GREEN = PatternFill("solid", fgColor="D9EAD3")    # hold
FILL_GREY = PatternFill("solid", fgColor="F2F2F2")     # neutral banding

STATUS_GREEN = PatternFill("solid", fgColor="38A169")
STATUS_AMBER = PatternFill("solid", fgColor="DD8B1A")
STATUS_RED = PatternFill("solid", fgColor="D64545")
STATUS_FONT = Font(bold=True, size=20, color="FFFFFF")

THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY_FMT = "#,##0"
PCT_FMT = '0.0"%"'

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


# --------------------------------------------------------------------------- #
# Small helpers (defensive column/value access)
# --------------------------------------------------------------------------- #
def _col(df: pd.DataFrame, *candidates):
    """Return the first candidate column name that exists in df, else None."""
    for c in candidates:
        if c in df.columns:
            return c
    return None


def _series(df: pd.DataFrame, *candidates, default=np.nan):
    """Return the first matching column as a Series, else a default-filled one."""
    name = _col(df, *candidates)
    if name is not None:
        return df[name]
    return pd.Series([default] * len(df), index=df.index)


def _num(x, default=0.0) -> float:
    """Coerce a scalar to float, tolerating None / NaN / strings."""
    try:
        if x is None:
            return default
        v = float(x)
        if np.isnan(v):
            return default
        return v
    except (TypeError, ValueError):
        return default


def _fmt_val(v):
    """Human-friendly rendering of a scorecard scalar for label/value rows."""
    if v is None:
        return "-"
    if isinstance(v, float):
        if np.isnan(v):
            return "-"
        # keep small ratios readable, big money-ish numbers grouped
        if abs(v) >= 1000:
            return f"{v:,.0f}"
        return f"{v:,.3f}".rstrip("0").rstrip(".")
    return v


def _action_from_bucket(bucket) -> str:
    """Map shared-contract bucket -> row action vocabulary (cut/reinvest/hold)."""
    b = str(bucket or "").lower()
    if "waste" in b or b == "c_waste_cut" or b.startswith("c_"):
        return "cut"
    if "reinvest" in b or b == "e_reinvest" or b.startswith("e_"):
        return "reinvest"
    return "hold"


def _action_label(action: str) -> str:
    """Pretty label for the Action cell."""
    return {"cut": "CUT", "reinvest": "REINVEST", "hold": "HOLD"}.get(
        str(action).lower(), str(action or "HOLD").upper()
    )


def _autosize(ws, max_width=52, min_width=9, per_col_min=None):
    """Approximate autosize: width = max cell string length (+pad), clamped."""
    per_col_min = per_col_min or {}
    for col_cells in ws.columns:
        letter = None
        longest = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            if letter is None:
                letter = cell.column_letter
            longest = max(longest, len(str(cell.value)))
        if letter is None:
            continue
        width = min(max_width, max(min_width, longest + 2))
        width = max(width, per_col_min.get(letter, 0))
        ws.column_dimensions[letter].width = width


def _style_header_row(ws, row_idx, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


# --------------------------------------------------------------------------- #
# Sheet 1 — Weekly Plan
# --------------------------------------------------------------------------- #
def _build_weekly_plan(ws, tracker_df: pd.DataFrame, week_label: str):
    """One row per SKU x city cell: what to do this week + why + weekly saving."""
    headers = [
        "Product", "City", "Category", "Current Price", "Current Disc%",
        "This-Week Disc%", "This-Week Price", "Action", "Confidence",
        "Reason", "Weekly Saving (INR)",
    ]

    # Title band
    ws.cell(row=1, column=1, value=f"Weekly Discount Plan — {week_label}").font = TITLE_FONT
    ws.cell(
        row=2, column=1,
        value="One row per product x city. Colour tells you the move: amber = cut, "
              "blue = reinvest, green = hold.",
    ).font = SUBTITLE_FONT
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(headers))
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=len(headers))

    header_row = 4
    for j, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=j, value=h)
    _style_header_row(ws, header_row, len(headers))

    # Resolve source columns once (defensive).
    df = tracker_df.reset_index(drop=True)
    s_product = _series(df, "title", "Product", "product", default="")
    s_city = _series(df, "city", "City", default="")
    s_cat = _series(df, "category", "Category", default="")
    s_cur_price = _series(df, "cur_price", "Current Price", default=np.nan)
    s_cur_disc = _series(df, "cur_disc", "Current Disc%", default=np.nan)
    # The guardrailed move (week_disc/week_price) is what the owner acts on — it
    # wins over the raw model suggestion, which hasn't passed the seatbelt yet.
    s_wk_disc = _series(df, "week_disc", "this_week_disc", "suggested_disc",
                        "This-Week Disc%", default=np.nan)
    s_wk_price = _series(df, "week_price", "this_week_price", "suggested_price",
                         "This-Week Price", default=np.nan)
    s_conf = _series(df, "confidence", "Confidence", default="")
    s_reason = _series(df, "decision_reason", "reason", "Reason", default="")

    # Action: explicit week_action wins; else derive from bucket.
    action_col = _col(df, "week_action", "action", "Action")
    bucket_col = _col(df, "bucket")

    # Weekly saving: explicit column wins; else derive from rev delta
    # (positive delta => saving/gain).
    saving_col = _col(df, "week_saving_inr", "weekly_saving", "net_gain_wk", "Weekly Saving (INR)")
    delta_col = _col(df, "pred_net_rev_delta_wk")

    r = header_row + 1
    total_saving = 0.0
    for i in range(len(df)):
        # Action resolution
        if action_col is not None:
            action = str(df.at[i, action_col]).lower()
        elif bucket_col is not None:
            action = _action_from_bucket(df.at[i, bucket_col])
        else:
            action = "hold"

        # Weekly saving resolution
        if saving_col is not None:
            saving = _num(df.at[i, saving_col])
        elif delta_col is not None:
            saving = _num(df.at[i, delta_col])
        else:
            saving = 0.0
        total_saving += saving

        row_vals = [
            s_product.iloc[i],
            s_city.iloc[i],
            s_cat.iloc[i],
            _num(s_cur_price.iloc[i]),
            _num(s_cur_disc.iloc[i]),
            _num(s_wk_disc.iloc[i]),
            _num(s_wk_price.iloc[i]),
            _action_label(action),
            s_conf.iloc[i] if str(s_conf.iloc[i]) != "nan" else "",
            s_reason.iloc[i] if str(s_reason.iloc[i]) != "nan" else "",
            saving,
        ]
        for j, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.border = BORDER
            cell.font = BODY_FONT

        # Number formats
        ws.cell(row=r, column=4).number_format = MONEY_FMT   # current price
        ws.cell(row=r, column=5).number_format = PCT_FMT     # current disc%
        ws.cell(row=r, column=6).number_format = PCT_FMT     # this-week disc%
        ws.cell(row=r, column=7).number_format = MONEY_FMT   # this-week price
        ws.cell(row=r, column=11).number_format = MONEY_FMT  # weekly saving

        # Alignments
        for j in (1, 3, 8, 9, 10):
            ws.cell(row=r, column=j).alignment = LEFT
        for j in (2,):
            ws.cell(row=r, column=j).alignment = CENTER
        for j in (4, 5, 6, 7, 11):
            ws.cell(row=r, column=j).alignment = RIGHT

        # Conditional row fill by action
        fill = {"cut": FILL_AMBER, "reinvest": FILL_BLUE, "hold": FILL_GREEN}.get(
            action, FILL_GREEN
        )
        for j in range(1, len(headers) + 1):
            ws.cell(row=r, column=j).fill = fill
        r += 1

    # Total row
    if len(df) > 0:
        ws.cell(row=r, column=10, value="TOTAL weekly saving").font = LABEL_FONT
        ws.cell(row=r, column=10).alignment = RIGHT
        tc = ws.cell(row=r, column=11, value=total_saving)
        tc.font = Font(bold=True, size=11, color="1F4E5F")
        tc.number_format = MONEY_FMT
        tc.alignment = RIGHT
        tc.border = BORDER
    else:
        ws.cell(row=r, column=1, value="No cells in this week's plan.").font = SUBTITLE_FONT

    # Freeze header + autosize
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    _autosize(
        ws, per_col_min={
            "A": 26, "B": 12, "C": 14, "D": 13, "E": 13,
            "F": 15, "G": 15, "H": 11, "I": 13, "J": 40, "K": 18,
        },
    )
    # Reason column can be long — cap and wrap instead of exploding width.
    ws.column_dimensions["J"].width = 44
    for rr in range(header_row + 1, r + 1):
        ws.cell(row=rr, column=10).alignment = LEFT_WRAP


# --------------------------------------------------------------------------- #
# Sheet 2 — Guardrail dashboard
# --------------------------------------------------------------------------- #
def _status_color(status: str):
    s = str(status or "").strip().upper()
    if s.startswith("GREEN") or s == "OK":
        return STATUS_GREEN
    if s.startswith("AMBER") or s.startswith("YELLOW") or s.startswith("WARN"):
        return STATUS_AMBER
    if s.startswith("RED") or s.startswith("OVER") or s.startswith("BREACH"):
        return STATUS_RED
    return STATUS_AMBER


def _derive_guardrail(tracker_df: pd.DataFrame, guardrail_summary: dict) -> dict:
    """Fill missing guardrail fields from tracker_df where possible."""
    g = dict(guardrail_summary or {})

    gross = _num(g.get("total_gross_wk", g.get("gross_sales_wk", g.get("gross_sales"))))
    spend = _num(g.get("total_disc_spend_wk", g.get("disc_spend_wk",
                        g.get("discount_spend_wk", g.get("discount_spend")))))

    # Derive from tracker_df if not provided.
    if gross <= 0:
        gross = _num(_series(tracker_df, "cur_net_rev_wk", default=0.0).sum())
    if spend <= 0:
        spend = _num(_series(tracker_df, "cur_disc_spend_wk", default=0.0).sum())

    disc_pct = _num(g.get("disc_pct", g.get("discount_pct")))
    if 0 < disc_pct <= 1.0:                 # guardrail emits a fraction; render as %
        disc_pct *= 100.0
    if disc_pct <= 0 and gross > 0:
        disc_pct = 100.0 * spend / gross

    cap_pct = _num(g.get("budget_cap_pct", g.get("budget_pct_cap",
                         g.get("cap_pct"))))
    # budget_pct_cap in config is a fraction (0.11) => convert to percent.
    if 0 < cap_pct <= 1.0:
        cap_pct *= 100.0

    headroom = g.get("headroom_inr", g.get("headroom"))
    if headroom is None and cap_pct > 0:
        headroom = gross * cap_pct / 100.0 - spend
    headroom = _num(headroom)

    status = g.get("status")
    if not status:
        if cap_pct <= 0:
            status = "AMBER"
        elif disc_pct <= cap_pct * 0.9:
            status = "GREEN"
        elif disc_pct <= cap_pct:
            status = "AMBER"
        else:
            status = "RED"

    g.update({
        "gross_sales_wk": gross, "disc_spend_wk": spend, "disc_pct": disc_pct,
        "budget_cap_pct": cap_pct, "headroom_inr": headroom, "status": status,
    })
    return g


def _build_guardrail(ws, tracker_df: pd.DataFrame, guardrail_summary: dict):
    g = _derive_guardrail(tracker_df, guardrail_summary)

    ws.cell(row=1, column=1, value="Guardrail — Budget Safety Check").font = TITLE_FONT
    ws.cell(
        row=2, column=1,
        value="Is total discount spend inside the budget cap? Watch the big STATUS box.",
    ).font = SUBTITLE_FONT
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=4)
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=4)

    metrics = [
        ("Gross sales / week (INR)", g["gross_sales_wk"], MONEY_FMT),
        ("Discount spend / week (INR)", g["disc_spend_wk"], MONEY_FMT),
        ("Discount %", g["disc_pct"], PCT_FMT),
        ("Budget cap %", g["budget_cap_pct"], PCT_FMT),
        ("Headroom (INR left under cap)", g["headroom_inr"], MONEY_FMT),
    ]
    r = 4
    for label, val, fmt in metrics:
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = LABEL_FONT
        lc.border = BORDER
        lc.alignment = LEFT
        lc.fill = FILL_GREY
        vc = ws.cell(row=r, column=2, value=_num(val))
        vc.number_format = fmt
        vc.font = BODY_FONT
        vc.border = BORDER
        vc.alignment = RIGHT
        r += 1

    # Big STATUS box
    r += 1
    ws.cell(row=r, column=1, value="STATUS").font = LABEL_FONT
    status = str(g["status"]).upper()
    ws.merge_cells(start_row=r, end_row=r + 1, start_column=2, end_column=4)
    sc = ws.cell(row=r, column=2, value=status)
    sc.fill = _status_color(status)
    sc.font = STATUS_FONT
    sc.alignment = CENTER
    for rr in (r, r + 1):
        for cc in (2, 3, 4):
            ws.cell(row=rr, column=cc).border = BORDER
    legend_row = r + 2
    ws.cell(
        row=legend_row, column=1,
        value="GREEN = safe (<90% of cap)   AMBER = near cap   RED = over cap, cut spend",
    ).font = SUBTITLE_FONT
    ws.merge_cells(start_row=legend_row, end_row=legend_row,
                   start_column=1, end_column=4)

    # By-category mini table
    cat_row = legend_row + 2
    ws.cell(row=cat_row, column=1, value="Discount spend by category").font = LABEL_FONT
    cat_hdr = cat_row + 1
    cat_headers = ["Category", "Disc spend / wk (INR)", "Disc %"]
    for j, h in enumerate(cat_headers, start=1):
        ws.cell(row=cat_hdr, column=j, value=h)
    _style_header_row(ws, cat_hdr, len(cat_headers))

    cat_table = _resolve_category_table(tracker_df, guardrail_summary)
    rr = cat_hdr + 1
    for _, row in cat_table.iterrows():
        ws.cell(row=rr, column=1, value=row["category"]).border = BORDER
        ws.cell(row=rr, column=1).alignment = LEFT
        c2 = ws.cell(row=rr, column=2, value=_num(row["disc_spend"]))
        c2.number_format = MONEY_FMT
        c2.alignment = RIGHT
        c2.border = BORDER
        c3 = ws.cell(row=rr, column=3, value=_num(row["disc_pct"]))
        c3.number_format = PCT_FMT
        c3.alignment = RIGHT
        c3.border = BORDER
        rr += 1
    if len(cat_table) == 0:
        ws.cell(row=rr, column=1, value="(no category data)").font = SUBTITLE_FONT

    _autosize(ws, per_col_min={"A": 34, "B": 24, "C": 12, "D": 12})


def _resolve_category_table(tracker_df: pd.DataFrame, guardrail_summary: dict):
    """Use guardrail_summary['by_category'] if present, else derive from tracker_df."""
    g = guardrail_summary or {}
    by_cat = g.get("by_category")
    if by_cat is not None:
        if isinstance(by_cat, pd.DataFrame) and len(by_cat) > 0:
            df = by_cat.copy()
            # normalize column names
            colmap = {}
            for cand in ("category", "Category"):
                if cand in df.columns:
                    colmap[cand] = "category"
            for cand in ("disc_spend", "discount_spend", "disc_spend_wk"):
                if cand in df.columns:
                    colmap[cand] = "disc_spend"
            for cand in ("disc_pct", "discount_pct"):
                if cand in df.columns:
                    colmap[cand] = "disc_pct"
            df = df.rename(columns=colmap)
            for need in ("category", "disc_spend", "disc_pct"):
                if need not in df.columns:
                    df[need] = np.nan if need != "category" else ""
            return df[["category", "disc_spend", "disc_pct"]]
        if isinstance(by_cat, (list, dict)):
            try:
                df = pd.DataFrame(by_cat)
                if "category" in df.columns:
                    return df
            except Exception:
                pass

    # Derive from tracker_df.
    if tracker_df is None or len(tracker_df) == 0:
        return pd.DataFrame(columns=["category", "disc_spend", "disc_pct"])

    df = tracker_df.copy()
    cat = _col(df, "category", "Category")
    if cat is None:
        return pd.DataFrame(columns=["category", "disc_spend", "disc_pct"])
    spend = _series(df, "cur_disc_spend_wk", default=0.0)
    rev = _series(df, "cur_net_rev_wk", default=0.0)
    tmp = pd.DataFrame({
        "category": df[cat].astype(str),
        "spend": pd.to_numeric(spend, errors="coerce").fillna(0.0),
        "rev": pd.to_numeric(rev, errors="coerce").fillna(0.0),
    })
    grp = tmp.groupby("category", as_index=False).sum()
    grp["disc_spend"] = grp["spend"]
    grp["disc_pct"] = np.where(grp["rev"] > 0, 100.0 * grp["spend"] / grp["rev"], 0.0)
    return grp[["category", "disc_spend", "disc_pct"]].sort_values(
        "disc_spend", ascending=False
    )


# --------------------------------------------------------------------------- #
# Sheet 3 — Accuracy Scorecard
# --------------------------------------------------------------------------- #
def _build_scorecard(ws, scorecard: dict):
    ws.cell(row=1, column=1, value="Accuracy Scorecard").font = TITLE_FONT
    ws.cell(
        row=2, column=1,
        value="How close last week's predictions were to what actually happened. "
              "Builds trust over time.",
    ).font = SUBTITLE_FONT
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=4)
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=4)

    scorecard = scorecard or {}
    weekly = scorecard.get("weekly")

    # First-run detection: no weekly rows AND no headline metrics.
    headline_keys = [
        ("Hit rate (calls that went right)", "hit_rate"),
        ("Predicted vs actual R2", "pred_vs_actual_r2"),
        ("Units error (MAPE)", "units_mape"),
        ("Revenue bias", "revenue_bias_inr"),
        ("Cumulative realized saving (INR)", "cumulative_realized_saving_inr"),
    ]
    has_headline = any(scorecard.get(k) is not None for _, k in headline_keys)
    has_weekly = weekly is not None and len(weekly) > 0
    first_run = int(scorecard.get("n_weeks_scored", 0) or 0) == 0   # nothing scored yet

    r = 4
    if first_run or (not has_headline and not has_weekly):
        msg = ws.cell(
            row=r, column=1,
            value="No actuals yet — fills in from week 2.",
        )
        msg.font = Font(bold=True, size=12, color="DD8B1A")
        ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=4)
        _autosize(ws, per_col_min={"A": 34, "B": 16})
        return

    # Headline metrics label/value block
    ws.cell(row=r, column=1, value="Headline metrics").font = LABEL_FONT
    r += 1
    for label, key in headline_keys:
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = LABEL_FONT
        lc.fill = FILL_GREY
        lc.border = BORDER
        lc.alignment = LEFT
        val = scorecard.get(key)
        fmt = None
        if key == "cumulative_realized_saving_inr":
            fmt = MONEY_FMT
        vc = ws.cell(row=r, column=2,
                     value=_num(val) if fmt else _fmt_val(val))
        if fmt:
            vc.number_format = fmt
        vc.font = BODY_FONT
        vc.border = BORDER
        vc.alignment = RIGHT
        r += 1

    # Weekly time series table
    r += 1
    ws.cell(row=r, column=1, value="Week-by-week").font = LABEL_FONT
    r += 1
    ts_headers = ["Week", "Hit rate", "Realized saving (INR)", "Pred-vs-actual R2"]
    for j, h in enumerate(ts_headers, start=1):
        ws.cell(row=r, column=j, value=h)
    _style_header_row(ws, r, len(ts_headers))
    r += 1

    ts = _normalize_weekly(weekly)
    if len(ts) == 0:
        ws.cell(row=r, column=1,
                value="No actuals yet — fills in from week 2.").font = SUBTITLE_FONT
    else:
        for _, row in ts.iterrows():
            ws.cell(row=r, column=1, value=row["week"]).border = BORDER
            ws.cell(row=r, column=1).alignment = CENTER
            c2 = ws.cell(row=r, column=2, value=_fmt_val(row["hit_rate"]))
            c2.alignment = RIGHT
            c2.border = BORDER
            c3 = ws.cell(row=r, column=3, value=_num(row["realized_saving"]))
            c3.number_format = MONEY_FMT
            c3.alignment = RIGHT
            c3.border = BORDER
            c4 = ws.cell(row=r, column=4, value=_fmt_val(row["r2"]))
            c4.alignment = RIGHT
            c4.border = BORDER
            r += 1

    _autosize(ws, per_col_min={"A": 34, "B": 16, "C": 22, "D": 18})


def _normalize_weekly(weekly):
    """Coerce weekly time series (DataFrame/list/dict) into a tidy DataFrame."""
    cols = ["week", "hit_rate", "realized_saving", "r2"]
    if weekly is None:
        return pd.DataFrame(columns=cols)
    if isinstance(weekly, pd.DataFrame):
        df = weekly.copy()
    else:
        try:
            df = pd.DataFrame(weekly)
        except Exception:
            return pd.DataFrame(columns=cols)
    if len(df) == 0:
        return pd.DataFrame(columns=cols)

    def pick(frame, *cands):
        for c in cands:
            if c in frame.columns:
                return frame[c]
        return pd.Series([np.nan] * len(frame), index=frame.index)

    out = pd.DataFrame({
        "week": pick(df, "week", "week_label", "Week").astype(str),
        "hit_rate": pick(df, "hit_rate", "Hit rate"),
        "realized_saving": pick(df, "realized_saving", "Realized saving",
                                "realized_saving_inr"),
        "r2": pick(df, "pred_vs_actual_r2", "r2", "Pred-vs-actual R2"),
    })
    return out


# --------------------------------------------------------------------------- #
# Sheet 4 — Seasonality
# --------------------------------------------------------------------------- #
def _build_seasonality(ws, season_info, week_label: str):
    ws.cell(row=1, column=1, value="Seasonality & Festival Calendar").font = TITLE_FONT
    ws.cell(
        row=2, column=1,
        value="Festivals lift demand — the tool nudges discounts up a little around them.",
    ).font = SUBTITLE_FONT
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=4)
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=4)

    season_info = season_info or {}

    # This week's status block
    r = 4
    ws.cell(row=r, column=1, value=f"This week ({week_label})").font = LABEL_FONT
    r += 1
    status_rows = [
        ("Season status", "Festival week" if season_info.get("active") else "Normal"),
        ("Active festival", season_info.get("festival_name") or "None"),
        ("Uplift applied", f"+{season_info.get('budget_uplift_pct', 0)*100:.0f}%"
                           if season_info.get("active") else "0%"),
        ("Note", season_info.get("note", season_info.get("message", ""))),
    ]
    for label, val in status_rows:
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = LABEL_FONT
        lc.fill = FILL_GREY
        lc.border = BORDER
        lc.alignment = LEFT
        if label == "Uplift applied" and val is not None and str(val) != "":
            # render as +X% if it's a fraction/number
            uv = _num(val, default=None) if val is not None else None
            if uv is not None:
                shown = f"+{uv*100:.0f}%" if 0 <= uv <= 1 else f"+{uv:.1f}%"
            else:
                shown = str(val)
        else:
            shown = "" if val is None else str(val)
        vc = ws.cell(row=r, column=2, value=shown)
        vc.font = BODY_FONT
        vc.border = BORDER
        vc.alignment = LEFT
        r += 1

    # Festival calendar table
    r += 1
    ws.cell(row=r, column=1, value="Festival calendar").font = LABEL_FONT
    r += 1
    cal_headers = ["Festival", "Date", "Window", "Suggested uplift"]
    for j, h in enumerate(cal_headers, start=1):
        ws.cell(row=r, column=j, value=h)
    _style_header_row(ws, r, len(cal_headers))
    r += 1

    calendar = _resolve_calendar(season_info)
    if len(calendar) == 0:
        ws.cell(row=r, column=1,
                value="(no festival calendar provided)").font = SUBTITLE_FONT
    else:
        for row in calendar:
            ws.cell(row=r, column=1, value=str(row.get("festival", ""))).border = BORDER
            ws.cell(row=r, column=1).alignment = LEFT
            ws.cell(row=r, column=2, value=str(row.get("date", ""))).border = BORDER
            ws.cell(row=r, column=2).alignment = CENTER
            ws.cell(row=r, column=3, value=str(row.get("window", ""))).border = BORDER
            ws.cell(row=r, column=3).alignment = CENTER
            up = row.get("uplift", row.get("uplift_pct", ""))
            uv = _num(up, default=None)
            if uv is not None:
                up_str = f"+{uv*100:.0f}%" if 0 <= uv <= 1 else f"+{uv:.1f}%"
            else:
                up_str = str(up)
            ws.cell(row=r, column=4, value=up_str).border = BORDER
            ws.cell(row=r, column=4).alignment = CENTER
            r += 1

    _autosize(ws, per_col_min={"A": 22, "B": 16, "C": 22, "D": 18})


def _resolve_calendar(season_info: dict):
    """Return a list of festival dicts from season_info, or a sensible default."""
    cal = season_info.get("calendar") or season_info.get("festival_calendar")
    if cal is not None:
        if isinstance(cal, pd.DataFrame):
            return cal.to_dict("records")
        if isinstance(cal, (list, tuple)):
            return list(cal)
        if isinstance(cal, dict):
            return [cal]
    # Default India CPG festival skeleton (dates approximate, editable by user).
    return [
        {"festival": "Republic Day", "date": "26 Jan", "window": "23-27 Jan", "uplift": 0.3},
        {"festival": "Holi", "date": "Mar", "window": "±3 days", "uplift": 0.3},
        {"festival": "Independence Day", "date": "15 Aug", "window": "12-16 Aug", "uplift": 0.3},
        {"festival": "Raksha Bandhan", "date": "Aug", "window": "±3 days", "uplift": 0.4},
        {"festival": "Diwali", "date": "Oct/Nov", "window": "±7 days", "uplift": 0.5},
        {"festival": "New Year", "date": "31 Dec", "window": "28 Dec-1 Jan", "uplift": 0.4},
    ]


# --------------------------------------------------------------------------- #
# Sheet 5 — How to use
# --------------------------------------------------------------------------- #
def _build_how_to_use(ws, week_label: str):
    ws.cell(row=1, column=1, value="How to use this workbook").font = TITLE_FONT
    ws.cell(
        row=2, column=1,
        value="Plain-English steps. No spreadsheets skills needed — just read top to bottom.",
    ).font = SUBTITLE_FONT
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=6)
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=6)

    steps = [
        ("1. Get this file each week",
         "Every Monday, run the weekly export from the tool. It produces this "
         "workbook, named with the week (e.g. this one is " + week_label + ")."),
        ("2. Open the 'Weekly Plan' sheet first",
         "Each row is one product in one city. The colour tells you the move at a "
         "glance: AMBER = cut the discount, BLUE = reinvest (push a bit more), "
         "GREEN = hold / no change."),
        ("3. Read the Action, then the Reason",
         "'Action' says what to do. 'Reason' says why in one line. 'Confidence' "
         "tells you how sure the tool is: High = act now, Experimental = try small, "
         "Low = watch only."),
        ("4. Apply the 'This-Week Disc%' in Blinkit",
         "Set each cell's discount to the 'This-Week Disc%' value. The "
         "'This-Week Price' is what the shopper will see. Only change what the "
         "plan tells you to."),
        ("5. Check the 'Guardrail' sheet — the red/green rule",
         "Look at the big STATUS box. GREEN = you are safely under budget, go ahead. "
         "AMBER = you are close to the cap, be careful. RED = you are over budget — "
         "cut some discounts before you apply the plan."),
        ("6. Watch the 'Accuracy Scorecard' grow",
         "From week 2 onward this sheet shows how close last week's predictions were "
         "to reality. A rising hit-rate and realized saving means the tool is earning "
         "its keep. Week 1 will say 'No actuals yet'."),
        ("7. Mind the 'Seasonality' sheet around festivals",
         "Near Diwali, Holi, etc. demand jumps, so the tool nudges discounts up a "
         "little. Nothing for you to do — just know why numbers move in festival weeks."),
        ("8. THE GOLDEN RULE",
         "If a cut loses sales for 2 weeks in a row, REVERT it — put the discount "
         "back to where it was. Protecting volume beats chasing a saving that costs "
         "you customers."),
    ]

    r = 4
    for head, body in steps:
        hc = ws.cell(row=r, column=1, value=head)
        hc.font = Font(bold=True, size=12, color="1F4E5F")
        ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=6)
        r += 1
        bc = ws.cell(row=r, column=1, value=body)
        bc.font = BODY_FONT
        bc.alignment = LEFT_WRAP
        ws.merge_cells(start_row=r, end_row=r + 1, start_column=1, end_column=6)
        # Highlight the golden rule
        if head.strip().startswith("8"):
            for rr in (r, r + 1):
                for cc in range(1, 7):
                    ws.cell(row=rr, column=cc).fill = FILL_AMBER
        r += 3

    # Fixed, readable column widths (this is a text sheet).
    for letter, w in zip("ABCDEF", (18, 18, 18, 18, 18, 18)):
        ws.column_dimensions[letter].width = w


# --------------------------------------------------------------------------- #
# Public entry point
# --------------------------------------------------------------------------- #
def build_workbook(tracker_df, guardrail_summary, scorecard, season_info,
                   out_path, week_label) -> str:
    """
    Build the polished multi-sheet weekly tracker workbook.

    Parameters
    ----------
    tracker_df : pandas.DataFrame
        One row per SKU x city cell (plan_df + view aliases). See module docstring.
    guardrail_summary : dict
        Budget-safety numbers. Missing fields are derived from tracker_df.
        Optional key 'by_category' (DataFrame or list of dicts) drives the mini table.
    scorecard : dict
        Headline accuracy metrics + optional 'weekly' time series. Empty/first-run
        renders a friendly 'No actuals yet' message.
    season_info : dict
        This week's season status + optional 'calendar' festival table.
    out_path : str
        Destination .xlsx path. Parent dir is created if needed.
    week_label : str
        e.g. 'W1'. Used in titles and the How-to-use text.

    Returns
    -------
    str : out_path (the file that was written).
    """
    if tracker_df is None:
        tracker_df = pd.DataFrame()
    if not isinstance(tracker_df, pd.DataFrame):
        tracker_df = pd.DataFrame(tracker_df)
    guardrail_summary = guardrail_summary or {}
    scorecard = scorecard or {}
    season_info = season_info or {}
    week_label = str(week_label or "W?")

    wb = Workbook()

    ws_plan = wb.active
    ws_plan.title = "Weekly Plan"
    _build_weekly_plan(ws_plan, tracker_df, week_label)

    _build_guardrail(wb.create_sheet("Guardrail"), tracker_df, guardrail_summary)
    _build_scorecard(wb.create_sheet("Accuracy Scorecard"), scorecard)
    _build_seasonality(wb.create_sheet("Seasonality"), season_info, week_label)
    _build_how_to_use(wb.create_sheet("How to use"), week_label)

    # Nice-to-have: default the view to the plan sheet, gridlines off on text sheets.
    ws_plan.sheet_view.showGridLines = True
    wb["How to use"].sheet_view.showGridLines = False

    # Ensure parent dir exists.
    parent = os.path.dirname(os.path.abspath(out_path))
    if parent and not os.path.isdir(parent):
        os.makedirs(parent, exist_ok=True)

    wb.save(out_path)
    return out_path


# --------------------------------------------------------------------------- #
# Smoke test
# --------------------------------------------------------------------------- #
if __name__ == "__main__":
    import sys

    # Tiny synthetic plan_df / config per the shared contract.
    plan = pd.DataFrame([
        {
            "cell_id": "c1", "product_id": "p1", "city": "Mumbai",
            "category": "Oil", "title": "Cold Pressed Groundnut Oil 1L",
            "mrp": 450.0, "cur_price": 405.0, "cur_disc": 10.0,
            "cur_units_wk": 120, "cur_net_rev_wk": 120 * 405.0,
            "cur_disc_spend_wk": 120 * 450.0 * 10.0 / 100,
            "suggested_disc": 6.0, "suggested_price": 450.0 * (1 - 6.0 / 100),
            "pred_units_wk": 118, "pred_net_rev_wk": 118 * 450.0 * 0.94,
            "pred_net_rev_delta_wk": 118 * 450.0 * 0.94 - 120 * 405.0,
            "bucket": "c_waste_cut", "confidence": "High",
            "reliably_waste": True, "net_gain_mo": 21000.0,
            "week_action": "cut", "weekly_saving": 5250.0,
            "decision_reason": "Discount not driving units — trim to recover margin.",
        },
        {
            "cell_id": "c2", "product_id": "p2", "city": "Delhi",
            "category": "Salt", "title": "Rock Salt 1kg",
            "mrp": 90.0, "cur_price": 81.0, "cur_disc": 10.0,
            "cur_units_wk": 300, "cur_net_rev_wk": 300 * 81.0,
            "cur_disc_spend_wk": 300 * 90.0 * 10.0 / 100,
            "suggested_disc": 14.0, "suggested_price": 90.0 * (1 - 14.0 / 100),
            "pred_units_wk": 360, "pred_net_rev_wk": 360 * 90.0 * 0.86,
            "pred_net_rev_delta_wk": 360 * 90.0 * 0.86 - 300 * 81.0,
            "bucket": "e_reinvest", "confidence": "Experimental",
            "reliably_waste": False, "net_gain_mo": 12000.0,
            "week_action": "reinvest", "weekly_saving": 3564.0,
            "decision_reason": "Elastic cell — deeper cut lifts volume profitably.",
        },
        {
            "cell_id": "c3", "product_id": "p3", "city": "Bengaluru",
            "category": "Spices", "title": "Turmeric Powder 200g",
            "mrp": 120.0, "cur_price": 108.0, "cur_disc": 10.0,
            "cur_units_wk": 200, "cur_net_rev_wk": 200 * 108.0,
            "cur_disc_spend_wk": 200 * 120.0 * 10.0 / 100,
            "suggested_disc": 10.0, "suggested_price": 108.0,
            "pred_units_wk": 200, "pred_net_rev_wk": 200 * 108.0,
            "pred_net_rev_delta_wk": 0.0,
            "bucket": "f_monitor", "confidence": "Low",
            "reliably_waste": False, "net_gain_mo": 0.0,
            "week_action": "hold", "weekly_saving": 0.0,
            "decision_reason": "Not enough signal — hold and monitor another week.",
        },
    ])

    config = {
        "budget_pct_cap": 0.11, "max_step_ppt": 3.0, "festival_uplift_pct": 0.5,
        "week_date": "2026-07-06", "week_label": "W1",
    }

    guardrail_summary = {
        "gross_sales_wk": float(plan["cur_net_rev_wk"].sum()),
        "disc_spend_wk": float(plan["cur_disc_spend_wk"].sum()),
        "budget_cap_pct": config["budget_pct_cap"],  # fraction -> auto-converted
    }

    # First-run scorecard (empty) to exercise the 'No actuals yet' path,
    # plus a populated variant to exercise the table path.
    scorecard_empty = {}
    scorecard_full = {
        "hit_rate": 0.82, "pred_vs_actual_r2": 0.79, "units_mape": 0.14,
        "revenue_bias": -0.02, "cumulative_realized_saving": 41250.0,
        "weekly": [
            {"week": "W1", "hit_rate": 0.80, "realized_saving": 18000.0,
             "pred_vs_actual_r2": 0.76},
            {"week": "W2", "hit_rate": 0.84, "realized_saving": 23250.0,
             "pred_vs_actual_r2": 0.79},
        ],
    }

    season_info = {
        "status": "Normal week (no active festival)",
        "festival": "None",
        "uplift_pct": 0.0,
        "note": "Next festival window: Independence Day (12-16 Aug).",
    }

    tmpdir = tempfile.gettempdir()
    out1 = os.path.join(tmpdir, "weekly_tracker_smoke_firstrun.xlsx")
    out2 = os.path.join(tmpdir, "weekly_tracker_smoke_populated.xlsx")

    p1 = build_workbook(plan, guardrail_summary, scorecard_empty, season_info,
                        out1, config["week_label"])
    p2 = build_workbook(plan, guardrail_summary, scorecard_full, season_info,
                        out2, config["week_label"])

    # Re-open both to confirm they are valid workbooks with the expected sheets.
    expected_sheets = ["Weekly Plan", "Guardrail", "Accuracy Scorecard",
                       "Seasonality", "How to use"]
    ok = True
    for p in (p1, p2):
        wb = load_workbook(p)
        sheets = wb.sheetnames
        missing = [s for s in expected_sheets if s not in sheets]
        plan_ws = wb["Weekly Plan"]
        frozen = plan_ws.freeze_panes
        print(f"[OK] wrote & reopened: {p}")
        print(f"     sheets: {sheets}")
        print(f"     Weekly Plan freeze_panes: {frozen}")
        print(f"     Weekly Plan dims: {plan_ws.max_row} rows x "
              f"{plan_ws.max_column} cols")
        if missing:
            ok = False
            print(f"     [FAIL] missing sheets: {missing}")

    # Spot-check a couple of values on the populated file.
    wb2 = load_workbook(out2)
    gr = wb2["Guardrail"]
    print("[CHECK] Guardrail gross-sales cell B4 =", gr["B4"].value)
    sc = wb2["Accuracy Scorecard"]
    print("[CHECK] Scorecard first label A5 =", sc["A5"].value)

    print("\nRESULT:", "PASS — all sheets present, files valid." if ok
          else "FAIL — see messages above.")
    sys.exit(0 if ok else 1)
