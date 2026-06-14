"""
Excel report generator — McKinsey-minimalist style with live formulas.

Design principles:
  - All numbers live on a hidden Data sheet
  - Summary, By Product, and detail sheets reference Data via SIMPLE
    formulas (SUM, SUMIF, SUMPRODUCT, IF, ROUND) — compatible with
    Excel 2010+ (no XLOOKUP, LET, LAMBDA, dynamic arrays, etc.)
  - The user can edit Data values and watch Summary recalculate live —
    nothing is hard-coded except section labels and the date stamp.
  - Styling: single font (Calibri), navy accent used sparingly,
    hairline borders only, lots of whitespace.

Sheets created:
  Summary       — Portfolio + This Week's Plan + Model Accuracy
  By Product    — Per-SKU breakdown
  Price Lifts   — Detailed cut recommendations
  Price Drops   — Detailed reinvest recommendations
  Needs Test    — Low-confidence cells
  Data          — Raw per-cell data (source of truth for formulas)
"""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import os


# ────────────────────────────────────────────────────────────────────
# Design tokens
# ────────────────────────────────────────────────────────────────────
FONT_FAMILY = "Calibri"
INK   = "FF0F172A"   # primary text (near-black)
BODY  = "FF1F2937"
MUTED = "FF6B7280"
ACCENT = "FF1E3A5F"  # slate navy — used very sparingly
POS    = "FF15803D"  # green: savings / volume gain
NEG    = "FFB91C1C"  # red: spend up / gap to target

THIN  = Side(border_style="thin",  color="FFE5E7EB")  # hairline
RULE  = Side(border_style="thin",  color="FF9CA3AF")  # medium rule
BOLD_RULE = Side(border_style="medium", color="FF0F172A")  # heavy rule

NO_FILL = PatternFill(fill_type=None)


def f(size=10, bold=False, color=BODY, italic=False):
    """Helvetica-equivalent (Calibri) font shortcut."""
    return Font(name=FONT_FAMILY, size=size, bold=bold, color=color, italic=italic)


def al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def b(top=None, bottom=None, left=None, right=None):
    return Border(top=top, bottom=bottom, left=left, right=right)


# ────────────────────────────────────────────────────────────────────
# Public entry point
# ────────────────────────────────────────────────────────────────────
def write_excel(summary, waste_main, reinvest_main, needs_test, df, run_dir):
    """
    Build the .xlsx report and return the file path.
    `summary` must include the 'business' and 'model_accuracy' sub-dicts.
    `df` is the enriched recommendations DataFrame (post Stage 8 enrich).
    """
    wb = Workbook()
    # Default sheet — rename and use as Summary
    ws_sum   = wb.active
    ws_sum.title = "Summary"
    ws_glide = wb.create_sheet("Glide Path")
    ws_track = wb.create_sheet("Track Record")
    ws_leak  = wb.create_sheet("Leakage")
    ws_prod  = wb.create_sheet("By Product")
    ws_cut   = wb.create_sheet("Price Lifts")
    ws_inv   = wb.create_sheet("Price Drops")
    ws_test  = wb.create_sheet("Needs Test")
    ws_data  = wb.create_sheet("Data")

    # 1. Build the Data sheet first — everything else references it.
    _build_data_sheet(ws_data, df, waste_main, reinvest_main)

    # 2. Build the Summary using formulas that reference Data!
    _build_summary_sheet(ws_sum, summary, df)

    # 2b. Multi-cycle roadmap — week-by-week projection
    _build_glide_path_sheet(ws_glide, summary)

    # 2c. Track Record — the receipts (out-of-time backtest + live results)
    _build_track_record_sheet(ws_track, summary)

    # 2d. Leakage — real vs borrowed vs stolen + the "worth discounting?" gate
    _build_leakage_sheet(ws_leak, summary)

    # 3. Per-product breakdown — city-by-city, with full weekly glide
    _build_per_product_sheet(ws_prod, summary, df, waste_main, reinvest_main)

    # 4. Detail sheets — straightforward tables
    _build_detail_sheet(ws_cut,  waste_main,    sheet_type="cut")
    _build_detail_sheet(ws_inv,  reinvest_main, sheet_type="invest")
    _build_detail_sheet(ws_test, needs_test,    sheet_type="needs_test")

    # Hide the raw data sheet — power users can unhide it
    ws_data.sheet_state = "hidden"

    path = os.path.join(run_dir, "WASTE_REINVEST_REPORT.xlsx")
    wb.save(path)
    return path


# ────────────────────────────────────────────────────────────────────
# Sheet builders
# ────────────────────────────────────────────────────────────────────
def _build_data_sheet(ws, df, waste_main, reinvest_main):
    """
    Raw per-cell data + the recommended price (cut or invest) for the week.
    Formulas elsewhere SUMPRODUCT off these columns.

    Columns: cell_id | product | grammage | city | mrp |
             cur_disc% | cur_units_day | cur_price |
             aftercut_disc% | aftercut_units | aftercut_price   (cuts only applied)
             final_disc% | final_units | final_price             (cuts + reinvest)
             confidence | elasticity | category
    """
    headers = [
        "cell_id", "product", "grammage", "city", "mrp",
        "cur_disc_pct", "cur_units_day", "cur_price",
        "aftercut_disc_pct", "aftercut_units", "aftercut_price",
        "final_disc_pct", "final_units", "final_price",
        "confidence", "elasticity", "category",
    ]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = f(10, bold=True, color=INK)
        c.alignment = al("center")
        c.border = b(bottom=RULE)

    # Map cell_id → plan price (cut)
    cut_map = {}
    if not waste_main.empty:
        for _, r in waste_main.iterrows():
            cid = r.get("cell_id")
            if pd.notna(cid):
                cut_map[cid] = {
                    "disc":  float(r.get("rec_discount_final", r["current_discount_pct"])),
                    "units": float(r.get("rec_units_day", r["current_units_day"]))
                                if "rec_units_day" in r and pd.notna(r.get("rec_units_day"))
                                else _predict(r, float(r.get("rec_discount_final",
                                                              r["current_discount_pct"]))),
                }
    inv_map = {}
    if not reinvest_main.empty:
        for _, r in reinvest_main.iterrows():
            cid = r.get("cell_id")
            if pd.notna(cid):
                rec_d = float(r.get("rec_discount_final",
                                      r.get("recommended_discount_pct",
                                            r["current_discount_pct"])))
                # ALWAYS re-predict for reinvest cells — `rec_units_day` in the
                # row was set by Stage 7's CUT logic (price-up), not the
                # reinvest action (price-down), so it's wrong here.
                inv_map[cid] = {"disc": rec_d, "units": _predict(r, rec_d)}

    rrow = 2
    for _, row in df.iterrows():
        cid = row.get("cell_id")
        mrp = float(row.get("mrp", 0))
        cur_d = float(row.get("current_discount_pct", 0))
        cur_u = float(row.get("current_units_day", 0))

        # AFTER CUTS only: apply cut to cells in cut list (waste_main is already
        # filtered to exclude reinvest cells, so 'in cut_map' implies 'not reinvest')
        if cid in cut_map:
            ac_d, ac_u = cut_map[cid]["disc"], cut_map[cid]["units"]
        else:
            ac_d, ac_u = cur_d, cur_u

        # FINAL (cuts + reinvest): reinvest takes precedence; then cuts; else current
        if cid in inv_map:
            fn_d, fn_u = inv_map[cid]["disc"], inv_map[cid]["units"]
        elif cid in cut_map:
            fn_d, fn_u = cut_map[cid]["disc"], cut_map[cid]["units"]
        else:
            fn_d, fn_u = cur_d, cur_u

        values = [
            cid, str(row.get("title", ""))[:50], row.get("grammage", ""), row.get("city", ""), mrp,
            cur_d, cur_u, mrp * (1 - cur_d / 100),
            ac_d, ac_u, mrp * (1 - ac_d / 100),
            fn_d, fn_u, mrp * (1 - fn_d / 100),
            row.get("confidence", ""),
            float(row.get("price_elasticity", row.get("elasticity", 0))),
            row.get("category", ""),
        ]
        for j, v in enumerate(values, 1):
            c = ws.cell(row=rrow, column=j, value=v)
            c.font = f(10)
            if j in (5, 8, 11, 14):
                c.number_format = "#,##0.00"
            elif j in (6, 9, 12):
                c.number_format = "0.00"
            elif j in (7, 10, 13):
                c.number_format = "#,##0.0"
            elif j == 16:
                c.number_format = "0.00"
        rrow += 1

    # Reasonable column widths
    widths = [22, 36, 8, 18, 9, 11, 13, 11, 11, 14, 11, 13, 16, 12, 14, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _predict(row, new_d):
    """Predict units at new_d using the dual-signal model — fallback when
    rec_units_day isn't already in the row."""
    try:
        import numpy as np
        mrp   = float(row["mrp"])
        cur_d = float(row["current_discount_pct"])
        cur_u = float(row.get("current_units_day", 0))
        if cur_u <= 0 or mrp <= 0:
            return cur_u
        cur_p = mrp * (1 - cur_d / 100)
        new_p = mrp * (1 - new_d / 100)
        elast = float(row.get("price_elasticity", row.get("elasticity", -1.5)))
        badge = float(row.get("badge_sensitivity", row.get("discount_sensitivity", 0)))
        mult  = (new_p / cur_p) ** elast * np.exp(badge * (new_d - cur_d))
        return max(cur_u * mult, 0.01)
    except Exception:
        return float(row.get("current_units_day", 0))


def _build_summary_sheet(ws, summary, df):
    """
    The Portfolio Summary — every number except labels comes from a formula
    referencing the Data sheet.

    Layout (columns):
      A: label
      B: Today
      C: After cuts
      D: After cuts + invest
      E: notes / explanation
    """
    biz = summary.get("business", {})
    acc = summary.get("model_accuracy", {})
    target = _get_target_disc(summary)
    n_data_rows = len(df)
    last_row = 1 + n_data_rows  # Data sheet last row

    # Helpers to write formula cells
    def cell(r, c, val, font=None, align=None, fmt=None, border=None, fill=None):
        x = ws.cell(row=r, column=c, value=val)
        x.font = font or f(10)
        if align: x.alignment = align
        if fmt:   x.number_format = fmt
        if border:x.border = border
        if fill:  x.fill = fill
        return x

    # Title block
    cell(1, 1, "DISCOUNT OPTIMISATION  ·  WEEKLY REPORT  ·  " +
         pd.Timestamp.now().strftime("%d %B %Y"),
         font=f(9, color=MUTED), align=al("left"))
    cell(3, 1, "Portfolio Summary", font=f(20, bold=True, color=INK))
    cell(4, 1, "24 Mantra Organic on Blinkit", font=f(10, color=BODY))

    # ── Portfolio metrics table — formulas reference Data sheet ────────
    # Row 7 = header, rows 8-12 = metrics
    HEAD_ROW = 7
    R_GROSS  = 8
    R_SPEND  = 9
    R_NET    = 10
    R_UNITS  = 11
    R_DISC   = 12
    R_TARGET = 14

    headers = ["Metric", "Today", "After cuts", "After cuts + invest"]
    for j, h in enumerate(headers, 1):
        cell(HEAD_ROW, j, h,
             font=f(10, bold=True, color=INK),
             align=al("right" if j > 1 else "left"),
             border=b(top=BOLD_RULE, bottom=RULE))

    # Reference ranges on the Data sheet
    # Data layout: E=mrp, F=cur_disc%, G=cur_units, I=plan_disc%, J=plan_units, L=inv_disc%, M=inv_units
    rng_mrp        = f"Data!E2:E{last_row}"
    rng_cur_disc   = f"Data!F2:F{last_row}"
    rng_cur_units  = f"Data!G2:G{last_row}"
    rng_plan_disc  = f"Data!I2:I{last_row}"
    rng_plan_units = f"Data!J2:J{last_row}"
    rng_inv_disc   = f"Data!L2:L{last_row}"
    rng_inv_units  = f"Data!M2:M{last_row}"

    # Monthly multiplier baked in (30 days)
    K = 30  # days per month for sales × units → monthly

    # Gross sales (at MRP) = SUMPRODUCT(MRP × units × 30)
    cell(R_GROSS, 1, "Gross sales / month (at MRP)", font=f(10), align=al("left"))
    cell(R_GROSS, 2, f"=SUMPRODUCT({rng_mrp},{rng_cur_units})*{K}",
         fmt='"Rs."#,##0', align=al("right"))
    cell(R_GROSS, 3, f"=SUMPRODUCT({rng_mrp},{rng_cur_units})*{K}",
         fmt='"Rs."#,##0', align=al("right"))
    # ↑ Gross stays same in cuts scenario because cuts use plan_units which is same when nothing changes
    # Wait actually no — when cells are cut their units drop. So 'after cuts' gross = SUM(MRP × plan_units)
    cell(R_GROSS, 3, f"=SUMPRODUCT({rng_mrp},{rng_plan_units})*{K}",
         fmt='"Rs."#,##0', align=al("right"))
    cell(R_GROSS, 4, f"=SUMPRODUCT({rng_mrp},{rng_inv_units})*{K}",
         fmt='"Rs."#,##0', align=al("right"))

    # Discount spend = SUMPRODUCT(MRP × disc%/100 × units × 30)
    cell(R_SPEND, 1, "Discount spend / month", font=f(10), align=al("left"))
    cell(R_SPEND, 2, f"=SUMPRODUCT({rng_mrp},{rng_cur_disc},{rng_cur_units})/100*{K}",
         fmt='"Rs."#,##0', align=al("right"))
    cell(R_SPEND, 3, f"=SUMPRODUCT({rng_mrp},{rng_plan_disc},{rng_plan_units})/100*{K}",
         fmt='"Rs."#,##0', align=al("right"))
    cell(R_SPEND, 4, f"=SUMPRODUCT({rng_mrp},{rng_inv_disc},{rng_inv_units})/100*{K}",
         fmt='"Rs."#,##0', align=al("right"))

    # Net revenue = Gross − Discount spend  (formula references the cells above)
    cell(R_NET, 1, "Net revenue / month", font=f(10), align=al("left"))
    cell(R_NET, 2, f"=B{R_GROSS}-B{R_SPEND}", fmt='"Rs."#,##0', align=al("right"))
    cell(R_NET, 3, f"=C{R_GROSS}-C{R_SPEND}", fmt='"Rs."#,##0', align=al("right"))
    cell(R_NET, 4, f"=D{R_GROSS}-D{R_SPEND}", fmt='"Rs."#,##0', align=al("right"))

    # Units sold / month
    cell(R_UNITS, 1, "Units sold / month", font=f(10), align=al("left"))
    cell(R_UNITS, 2, f"=SUM({rng_cur_units})*{K}",  fmt="#,##0", align=al("right"))
    cell(R_UNITS, 3, f"=SUM({rng_plan_units})*{K}", fmt="#,##0", align=al("right"))
    cell(R_UNITS, 4, f"=SUM({rng_inv_units})*{K}",  fmt="#,##0", align=al("right"))

    # Weighted discount % = discount spend ÷ gross sales × 100   (the key business formula)
    cell(R_DISC, 1, "Weighted discount %",
         font=f(10, bold=True, color=INK), align=al("left"),
         border=b(top=RULE, bottom=BOLD_RULE))
    cell(R_DISC, 2, f"=B{R_SPEND}/B{R_GROSS}*100",
         fmt="0.00\"%\"", font=f(11, bold=True, color=INK), align=al("right"),
         border=b(top=RULE, bottom=BOLD_RULE))
    cell(R_DISC, 3, f"=C{R_SPEND}/C{R_GROSS}*100",
         fmt="0.00\"%\"", font=f(11, bold=True, color=INK), align=al("right"),
         border=b(top=RULE, bottom=BOLD_RULE))
    cell(R_DISC, 4, f"=D{R_SPEND}/D{R_GROSS}*100",
         fmt="0.00\"%\"", font=f(11, bold=True, color=INK), align=al("right"),
         border=b(top=RULE, bottom=BOLD_RULE))

    # Target + gap line
    cell(R_TARGET, 1, "Target weighted discount %",
         font=f(10, italic=True, color=MUTED), align=al("left"))
    cell(R_TARGET, 2, target, fmt="0.00\"%\"",
         font=f(10, italic=True, color=MUTED), align=al("right"))
    cell(R_TARGET + 1, 1, "Gap to target (today)",
         font=f(10, italic=True, color=MUTED), align=al("left"))
    cell(R_TARGET + 1, 2, f"=B{R_DISC}-B{R_TARGET}",
         fmt="+0.00\" ppt\";-0.00\" ppt\"",
         font=f(10, italic=True, bold=True, color=NEG), align=al("right"))
    cell(R_TARGET + 2, 1, "Gap to target (after this-week plan)",
         font=f(10, italic=True, color=MUTED), align=al("left"))
    cell(R_TARGET + 2, 2, f"=D{R_DISC}-B{R_TARGET}",
         fmt="+0.00\" ppt\";-0.00\" ppt\"",
         font=f(10, italic=True, color=MUTED), align=al("right"))

    # ── This Week's Plan table ────────────────────────────────────────
    P_TITLE = 19
    P_HEAD  = 21
    P_CUT   = 22
    P_INV   = 23
    P_NET   = 24

    # Counts already pre-computed in summary (waste_main is filtered to exclude reinvest cells)
    n_waste = int(biz.get("n_waste_cells", 0))
    n_reinv = int(biz.get("n_reinvest_cells", 0))

    cell(P_TITLE, 1, "This week's plan",
         font=f(14, bold=True, color=INK), align=al("left"))

    for j, h in enumerate(["Action", "Cells", "Discount spend Δ", "Units Δ / month"], 1):
        cell(P_HEAD, j, h,
             font=f(10, bold=True, color=INK),
             align=al("right" if j > 1 else "left"),
             border=b(top=BOLD_RULE, bottom=RULE))

    # Cut row: spend change = today_spend - after_cuts_spend
    cell(P_CUT, 1, "Cut (raise price)",  font=f(10), align=al("left"))
    cell(P_CUT, 2, n_waste, fmt="0", align=al("right"))
    cell(P_CUT, 3, f"=-(B{R_SPEND}-C{R_SPEND})", fmt='"Rs."#,##0;[Red]"Rs."-#,##0',
         font=f(10, color=POS), align=al("right"))
    cell(P_CUT, 4, f"=C{R_UNITS}-B{R_UNITS}", fmt="+#,##0;-#,##0",
         font=f(10, color=NEG), align=al("right"))

    # Reinvest row: spend change = after_cuts+invest_spend - after_cuts_spend
    cell(P_INV, 1, "Reinvest (drop price)", font=f(10), align=al("left"))
    cell(P_INV, 2, n_reinv, fmt="0", align=al("right"))
    cell(P_INV, 3, f"=D{R_SPEND}-C{R_SPEND}", fmt='"Rs."+#,##0;[Red]"Rs."-#,##0',
         font=f(10, color=NEG), align=al("right"))
    cell(P_INV, 4, f"=D{R_UNITS}-C{R_UNITS}", fmt="+#,##0;-#,##0",
         font=f(10, color=POS), align=al("right"))

    # Net row
    cell(P_NET, 1, "Net change", font=f(10, bold=True, color=INK), align=al("left"),
         border=b(top=RULE, bottom=BOLD_RULE))
    cell(P_NET, 2, f"=B{P_CUT}+B{P_INV}", fmt="0", align=al("right"),
         font=f(10, bold=True), border=b(top=RULE, bottom=BOLD_RULE))
    cell(P_NET, 3, f"=C{P_CUT}+C{P_INV}", fmt='"Rs."+#,##0;[Red]"Rs."-#,##0',
         font=f(10, bold=True, color=POS), align=al("right"),
         border=b(top=RULE, bottom=BOLD_RULE))
    cell(P_NET, 4, f"=D{P_CUT}+D{P_INV}", fmt="+#,##0;-#,##0",
         font=f(10, bold=True, color=NEG), align=al("right"),
         border=b(top=RULE, bottom=BOLD_RULE))

    # ── Model Accuracy table ──────────────────────────────────────────
    M_TITLE = 27
    M_HEAD  = 29
    M_R2    = 30
    M_MAPE  = 31
    M_TRAIN = 32
    M_TIER  = 34

    cell(M_TITLE, 1, "Model accuracy",
         font=f(14, bold=True, color=INK), align=al("left"))

    for j, h in enumerate(["Metric", "Value", "What it means"], 1):
        cell(M_HEAD, j, h,
             font=f(10, bold=True, color=INK),
             align=al("right" if j == 2 else "left"),
             border=b(top=BOLD_RULE, bottom=RULE))

    if acc.get("available"):
        cell(M_R2,    1, "Price-engine accuracy — held-out R²", font=f(10), align=al("left"))
        cell(M_R2,    2, acc.get("decision_r2_bin", acc.get("test_r2_agg", 0)),
             fmt="0.00", align=al("right"))
        cell(M_R2,    3, "Accuracy of the actual price/volume curve that sets every "
                          "recommendation (price effect only — no momentum shortcuts), on "
                          "data the model never saw. This is the number that governs the "
                          "plan. 1.0 = perfect, 0 = useless.",
             font=f(9, color=MUTED), align=al("left", wrap=True))

        cell(M_MAPE,  1, "Price-engine avg error (held-out, bin grain)", font=f(10), align=al("left"))
        cell(M_MAPE,  2, acc.get("decision_mape_bin", acc.get("test_mape_agg", 0)),
             fmt='0.0"%"', align=al("right"))
        cell(M_MAPE,  3, "Average % error of the price engine when comparing predicted vs "
                          "actual mean units in each 3-ppt discount band — the grain the "
                          "recommendation uses.",
             font=f(9, color=MUTED), align=al("left", wrap=True))

        cell(M_TRAIN, 1, "Full statistical model — held-out R² (context)", font=f(10), align=al("left"))
        cell(M_TRAIN, 2, acc["test_r2_log"], fmt="0.00", align=al("right"))
        cell(M_TRAIN, 3, "The broader model also uses recent-sales momentum & seasonality, "
                          "so its R² is higher — but that extra fit does NOT set prices. "
                          "Shown for context; the price-engine number above is the honest "
                          "one to quote.",
             font=f(9, color=MUTED), align=al("left", wrap=True))

        # Set tall row heights so wrapped text breathes
        for r in (M_R2, M_MAPE, M_TRAIN):
            ws.row_dimensions[r].height = 32

        # Overall tier — implemented as a nested IF formula so the bar is editable.
        cell(M_TIER, 1, "Overall tier",
             font=f(10, bold=True, color=INK), align=al("left"),
             border=b(top=RULE))
        # Tier formula: based on B30 (test R^2) and B31 (MAPE)
        tier_formula = (
            f'=IF(AND(B{M_R2}>=0.7,B{M_MAPE}<=25),"Strong",'
            f'IF(AND(B{M_R2}>=0.4,B{M_MAPE}<=50),"Moderate",'
            f'IF(AND(B{M_R2}>=0.1,B{M_MAPE}<=80),"Weak","Unreliable")))'
        )
        cell(M_TIER, 2, tier_formula,
             font=f(11, bold=True, color=ACCENT), align=al("right"),
             border=b(top=RULE))
        # Tier criteria reminder
        cell(M_TIER + 1, 1,
             "Strong: R²≥0.70 AND MAPE≤25%  |  Moderate: R²≥0.40 AND MAPE≤50%  |  "
             "Weak: R²≥0.10 AND MAPE≤80%  |  else Unreliable",
             font=f(8, italic=True, color=MUTED), align=al("left"))
        ws.merge_cells(start_row=M_TIER + 1, start_column=1,
                       end_row=M_TIER + 1, end_column=4)

        cell(M_TIER + 3, 1,
             f"Trained on {acc['n_train']:,} regular-day rows, validated on "
             f"{acc['n_test']:,} held-out future rows. Daily-level predictions are "
             f"inherently noisy for CPG SKU × city data; recommendations are most "
             f"reliable on the High-confidence cells.",
             font=f(9, italic=True, color=MUTED), align=al("left", wrap=True))
        ws.merge_cells(start_row=M_TIER + 3, start_column=1,
                       end_row=M_TIER + 3, end_column=5)
        ws.row_dimensions[M_TIER + 3].height = 32

    # ── Footer note ───────────────────────────────────────────────────
    F = M_TIER + 6
    cell(F, 1,
         "Discount % is computed as total monthly discount spend ÷ total monthly "
         "gross sales at MRP. Numbers in the Summary table are live formulas — "
         "edit the hidden Data sheet (Format ▸ Sheet ▸ Unhide) to run what-if scenarios.",
         font=f(8, color=MUTED), align=al("left", wrap=True))
    ws.merge_cells(start_row=F, start_column=1, end_row=F, end_column=5)
    ws.row_dimensions[F].height = 28

    # Column widths
    widths = [44, 22, 22, 24, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_glide_path_sheet(ws, summary):
    """
    Week-by-week roadmap: shows the user the full glide path the system
    has plotted from today to the target weighted discount.

    Columns:
       Cycle | Label | Weighted Disc % | Gross Sales | Discount Spend |
       Net Revenue | Units | Cumulative Savings | Gap to Target | Status

    User can change TARGET_TIMELINE_WEEKS in v4_config.py and re-run
    the pipeline to recompute this whole table.
    """
    glide = summary.get("glide_path")
    if glide is None or glide.empty:
        return

    def cell(r, c, val, font=None, align=None, fmt=None, border=None):
        x = ws.cell(row=r, column=c, value=val)
        x.font = font or f(10)
        if align: x.alignment = align
        if fmt: x.number_format = fmt
        if border: x.border = border
        return x

    target = _get_target_disc(summary)
    n_cycles = len(glide) - 1
    final_disc = float(glide.iloc[-1]["weighted_discount_pct"])
    final_gap  = final_disc - target
    cum_savings = float(glide.iloc[-1]["cumulative_savings"])

    # ── Header block ──────────────────────────────────────────────────
    cell(1, 1, "DISCOUNT OPTIMISATION  ·  GLIDE PATH (week-by-week roadmap)",
         font=f(9, color=MUTED), align=al("left"))
    cell(3, 1, "Multi-Cycle Roadmap", font=f(20, bold=True, color=INK))
    plan_max = getattr(__import__("v4_config"), "TARGET_TIMELINE_WEEKS", n_cycles)
    cell(4, 1, f"From {glide.iloc[0]['weighted_discount_pct']:.2f}% weighted discount today "
               f"to a projected {final_disc:.2f}% in {n_cycles} weekly cycles "
               f"(~{n_cycles/4.3:.1f} months). Budget was {plan_max} weeks — every "
               f"cell reached its target faster because gaps were small. "
               f"Per-cell endpoint = historical floor (the discount each cell "
               f"has proven it can survive).",
         font=f(10, color=BODY), align=al("left", wrap=True))
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=9)
    ws.row_dimensions[4].height = 26

    # Note on how the duration is set
    cell(6, 1,
         f"Each cell walks toward its individual target at a rate of "
         f"(current discount − target) ÷ {n_cycles} weeks. The portfolio's "
         f"weighted discount drops faster in early weeks (when the most "
         f"over-discounted cells move fastest) and slows as cells reach their "
         f"targets. To change the speed: edit TARGET_TIMELINE_WEEKS in v4_config.py.",
         font=f(9, italic=True, color=MUTED), align=al("left", wrap=True))
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=9)
    ws.row_dimensions[6].height = 36

    # Why the projection might not reach the 9% target
    biz = summary.get("business", {})
    n_actionable = int(biz.get("n_waste_cells", 0)) + int(biz.get("n_reinvest_cells", 0))
    needs_test_count = 33 - n_actionable  # rough — 33 is the typical portfolio size
    if final_gap > 0.5:
        cell(7, 1,
             f"This roadmap only moves the {n_actionable} cells the model is "
             f"confident enough to act on. The remaining cells are flagged "
             f"'Needs Price Test' (Sheet 'Needs Test') — they don't have enough clean "
             f"data to project safely. To close the rest of the gap to {target:.0f}%, "
             f"run small A/B price tests on those cells, add the test results to your "
             f"weekly data, and they'll join future runs of this glide path automatically.",
             font=f(9, italic=True, color=NEG), align=al("left", wrap=True))
        ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=9)
        ws.row_dimensions[7].height = 42

    # ── Headline outcome cards ────────────────────────────────────────
    H_ROW = 9
    cell(H_ROW, 1, "End-of-roadmap projection",
         font=f(12, bold=True, color=INK), align=al("left"))
    ws.merge_cells(start_row=H_ROW, start_column=1, end_row=H_ROW, end_column=9)

    outcome_data = [
        ["", "Today", f"After {n_cycles} weeks", "Change"],
        ["Weighted discount %",
            f"{glide.iloc[0]['weighted_discount_pct']:.2f}%",
            f"{final_disc:.2f}%",
            f"{final_disc - float(glide.iloc[0]['weighted_discount_pct']):+.2f} ppt"],
        ["Monthly discount spend",
            f"Rs. {float(glide.iloc[0]['discount_spend_inr']):,.0f}",
            f"Rs. {float(glide.iloc[-1]['discount_spend_inr']):,.0f}",
            f"-Rs. {cum_savings:,.0f}/mo"],
        ["Monthly net revenue",
            f"Rs. {float(glide.iloc[0]['net_revenue_inr']):,.0f}",
            f"Rs. {float(glide.iloc[-1]['net_revenue_inr']):,.0f}",
            f"+Rs. {float(glide.iloc[-1]['net_revenue_inr']) - float(glide.iloc[0]['net_revenue_inr']):,.0f}/mo"],
        ["Gap to target",
            f"{float(glide.iloc[0]['weighted_discount_pct']) - target:+.2f} ppt",
            f"{final_gap:+.2f} ppt",
            "TARGET REACHED" if final_gap <= 0.05 else f"still {final_gap:.2f} ppt short"],
    ]
    O_HEAD = H_ROW + 2
    for j, h in enumerate(outcome_data[0], 1):
        cell(O_HEAD, j, h, font=f(10, bold=True, color=INK),
             align=al("right" if j > 1 else "left"),
             border=b(top=BOLD_RULE, bottom=RULE))
    for i, row in enumerate(outcome_data[1:], 1):
        rr = O_HEAD + i
        for j, v in enumerate(row, 1):
            txt_color = INK
            if j == 4 and i == 1:  # weighted disc change
                txt_color = POS if "−" in str(v) or v.startswith("-") else MUTED
            elif j == 4 and i == 2:  # spend change
                txt_color = POS
            elif j == 4 and i == 3:  # revenue change
                txt_color = POS
            elif j == 4 and i == 4:
                txt_color = POS if "TARGET" in str(v) else NEG
            cell(rr, j, v, font=f(10, bold=(i == 4), color=txt_color),
                 align=al("right" if j > 1 else "left"))

    # ── The full week-by-week table ───────────────────────────────────
    T_TITLE = O_HEAD + 7
    cell(T_TITLE, 1, "Week-by-week projection",
         font=f(12, bold=True, color=INK), align=al("left"))

    T_HEAD = T_TITLE + 2
    headers = ["Cycle", "Label", "Weighted Disc %", "Gross Sales / mo",
               "Discount Spend / mo", "Net Revenue / mo", "Units / mo",
               "Cumulative Savings", "Gap to Target"]
    for j, h in enumerate(headers, 1):
        cell(T_HEAD, j, h, font=f(10, bold=True, color=INK),
             align=al("right" if j > 1 else "left"),
             border=b(top=BOLD_RULE, bottom=RULE))

    for i, row in glide.iterrows():
        rr = T_HEAD + 1 + i
        # Highlight the row where target is first reached
        is_first_target = bool(row["reached_target"]) and (
            i == 0 or not glide.iloc[i - 1]["reached_target"])
        bold = is_first_target
        bg_color = INK if bold else None
        # Today row gets bold treatment too
        is_today = (i == 0)
        if is_today or is_first_target:
            font_kwargs = {"bold": True, "color": INK}
        else:
            font_kwargs = {"color": BODY}

        cell(rr, 1, int(row["cycle"]), font=f(10, **font_kwargs),
             align=al("center"), fmt="0")
        cell(rr, 2, row["label"], font=f(10, **font_kwargs), align=al("left"))
        cell(rr, 3, float(row["weighted_discount_pct"]),
             font=f(10, **font_kwargs), align=al("right"), fmt="0.00\"%\"")
        cell(rr, 4, float(row["gross_sales_inr"]),
             font=f(10, **font_kwargs), align=al("right"), fmt="\"Rs.\"#,##0")
        cell(rr, 5, float(row["discount_spend_inr"]),
             font=f(10, **font_kwargs), align=al("right"), fmt="\"Rs.\"#,##0")
        cell(rr, 6, float(row["net_revenue_inr"]),
             font=f(10, **font_kwargs), align=al("right"), fmt="\"Rs.\"#,##0")
        cell(rr, 7, float(row["total_units"]),
             font=f(10, **font_kwargs), align=al("right"), fmt="#,##0")
        sav = float(row["cumulative_savings"])
        sav_color = POS if sav > 0 else MUTED
        cell(rr, 8, sav, font=f(10, color=sav_color, bold=font_kwargs.get("bold", False)),
             align=al("right"), fmt="\"Rs.\"+#,##0;\"Rs.\"-#,##0;-")
        gap = float(row["gap_to_target_ppt"])
        gap_color = NEG if gap > 0.5 else POS
        cell(rr, 9, gap, font=f(10, color=gap_color, bold=font_kwargs.get("bold", False)),
             align=al("right"), fmt="+0.00\" ppt\";-0.00\" ppt\";\"0.00 ppt\"")

        if is_first_target:
            for c in range(1, 10):
                ws.cell(row=rr, column=c).border = b(top=RULE, bottom=RULE)

    # Bottom border on last row
    last_data_row = T_HEAD + len(glide)
    for c in range(1, 10):
        ws.cell(row=last_data_row, column=c).border = b(bottom=BOLD_RULE)

    # Column widths
    widths = [8, 12, 17, 20, 20, 20, 15, 22, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{T_HEAD + 1}"


def _build_track_record_sheet(ws, summary):
    """
    Track Record — the receipts. Two parts:
      A. Out-of-time backtest: train as-of N weeks ago, grade forecasts on the
         weeks the model never saw. Proves it works before a brand acts.
      B. Live results: predicted vs realised once a brand acts (populates over
         time; shows a clear placeholder until then).

    Purely discount/volume based — no COGS or margin assumptions.
    """
    tr = summary.get("track_record") or {}
    bt = tr.get("backtest") or {}
    live = tr.get("live") or {}

    for col, w in {"A": 30, "B": 12, "C": 14, "D": 14, "E": 14, "F": 14, "G": 12}.items():
        ws.column_dimensions[col].width = w
    ws.sheet_view.showGridLines = False

    def cell(r, c, val, font=None, align=None, fmt=None, border=None, fill=None):
        x = ws.cell(row=r, column=c, value=val)
        x.font = font or f(10)
        if align: x.alignment = align
        if fmt: x.number_format = fmt
        if border: x.border = border
        if fill: x.fill = fill
        return x

    cell(1, 1, "DISCOUNT OPTIMISATION  ·  TRACK RECORD (proof the engine works)",
         font=f(9, color=MUTED), align=al("left"))
    cell(3, 1, "Track Record", font=f(20, bold=True, color=INK))

    if not bt.get("available"):
        cell(5, 1, "Backtest not available for this run — "
                   f"{bt.get('reason', 'insufficient data')}. "
                   "Run scripts/diagnostics/proof_loop.py once more data is in.",
             font=f(10, italic=True, color=MUTED), align=al("left", wrap=True))
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=7)
        return

    cell(4, 1, f"The engine was trained only on data up to {bt['cutoff']}, then asked to "
               f"predict the {bt['weeks']} weeks after (through {bt['max_date']}) — a window "
               f"it never saw. This is the honest test of whether following the tool "
               f"actually plays out. All figures are discount/volume only.",
         font=f(10, color=BODY), align=al("left", wrap=True))
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=7)
    ws.row_dimensions[4].height = 42

    # ── A. Forecast accuracy ──────────────────────────────────────────
    r = 7
    cell(r, 1, "A.  Out-of-time forecast accuracy", font=f(13, bold=True, color=INK))
    r += 1
    for j, h in enumerate(["Grain", "R²", "Avg error (MAPE)"], 1):
        cell(r, j, h, font=f(10, bold=True, color=INK),
             align=al("right" if j > 1 else "left"), border=b(top=BOLD_RULE, bottom=RULE))
    r += 1
    cell(r, 1, "Daily", align=al("left")); cell(r, 2, bt["daily_r2"], fmt="0.00", align=al("right"))
    cell(r, 3, bt["daily_mape"], fmt='0.0"%"', align=al("right")); r += 1
    if bt.get("bin_r2") is not None:
        cell(r, 1, "3-ppt discount bin", align=al("left"))
        cell(r, 2, bt["bin_r2"], fmt="0.00", align=al("right"))
        cell(r, 3, bt["bin_mape"], fmt='0.0"%"', align=al("right")); r += 1
    cell(r, 1, "A low forward R² is expected and OK — it means the engine doesn't predict the "
               "absolute future sales level (trend & seasonality dominate that). It is a "
               "price-RESPONSE model, not a demand forecaster. The real test is Part A2 below.",
         font=f(9, italic=True, color=MUTED), align=al("left", wrap=True))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.row_dimensions[r].height = 40
    r += 2

    # ── A2. Discount-move validation ──────────────────────────────────
    cell(r, 1, "A2.  Did volume move as predicted when price moved?",
         font=f(13, bold=True, color=INK)); r += 1
    for j, h in enumerate(["Forward discount move", "Cells", "Predicted vol Δ", "Actual vol Δ"], 1):
        cell(r, j, h, font=f(10, bold=True, color=INK),
             align=al("right" if j > 1 else "left"), border=b(top=BOLD_RULE, bottom=RULE))
    r += 1
    for row in bt.get("move_table", []):
        cell(r, 1, row["move"], align=al("left"))
        cell(r, 2, row["cells"], align=al("right"))
        cell(r, 3, row["pred"] / 100.0, fmt='+0.0%;-0.0%', align=al("right"))
        cell(r, 4, row["actual"] / 100.0, fmt='+0.0%;-0.0%', align=al("right"))
        r += 1
    r += 1

    # ── Verdict box ───────────────────────────────────────────────────
    vcolor = {"directional_conservative": POS, "directional_aggressive": ACCENT,
              "weak": NEG}.get(bt.get("verdict"), MUTED)
    n_cut = (bt.get("cut") or {}).get("cells")
    cut_sfx = f" · {n_cut} cells, 1 window" if n_cut else ""
    cell(r, 1, "Verdict", font=f(11, bold=True, color=INK), align=al("left"),
         border=b(top=RULE))
    cell(r, 2, {"directional_conservative": f"Conservative (directional{cut_sfx})",
                "directional_aggressive": f"Optimistic — quote savings low{cut_sfx}",
                "weak": "Inconclusive — too few clean moves"}.get(bt.get("verdict"), "—"),
         font=f(11, bold=True, color=vcolor), align=al("left"), border=b(top=RULE))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    r += 1
    cell(r, 1, bt.get("verdict_text", ""), font=f(10, color=BODY), align=al("left", wrap=True))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.row_dimensions[r].height = 56
    r += 2

    # ── B. Live results ───────────────────────────────────────────────
    cell(r, 1, "B.  Live results — predicted vs. actual, per city",
         font=f(13, bold=True, color=INK)); r += 1
    # Honest banner first
    cell(r, 1, live.get("note", ""), font=f(9, italic=True, color=MUTED),
         align=al("left", wrap=True))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.row_dimensions[r].height = 56
    r += 1
    if live.get("available") and live.get("cells"):
        if live.get("illustrative"):
            cell(r, 1, "↓ Illustrative (back-cast on the holdout — not a real acted cycle)",
                 font=f(9, bold=True, color=ACCENT), align=al("left")); r += 1
            cell(r, 1, "Per-city numbers are noisy — the model nails some cells and misses "
                       "others (shown honestly, not cherry-picked). The reliable signal is "
                       "the aggregate DIRECTION in Part A2, not any single city's figure.",
                 font=f(9, italic=True, color=MUTED), align=al("left", wrap=True))
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            ws.row_dimensions[r].height = 30; r += 1
        hdr = ["Product · City", "Price was", "Price became",
               "Predicted units", "Actual units", "Predicted vol Δ", "Actual vol Δ"]
        for j, h in enumerate(hdr, 1):
            cell(r, j, h, font=f(9, bold=True, color=INK),
                 align=al("right" if j > 1 else "left"), border=b(top=BOLD_RULE, bottom=RULE))
        r += 1
        for c in live["cells"]:
            cell(r, 1, c.get("label", ""), align=al("left"))
            cell(r, 2, c.get("base_price"), fmt='"Rs."#,##0', align=al("right"))
            cell(r, 3, c.get("achieved_price"), fmt='"Rs."#,##0', align=al("right"))
            cell(r, 4, c.get("pred_units"), fmt="#,##0", align=al("right"))
            cell(r, 5, c.get("actual_units"), fmt="#,##0", align=al("right"))
            cell(r, 6, (c.get("pred_vol") or 0) / 100.0, fmt='+0.0%;-0.0%', align=al("right"))
            cell(r, 7, (c.get("actual_vol") or 0) / 100.0, fmt='+0.0%;-0.0%', align=al("right"))
            r += 1
        r += 1

    cell(r, 1, "What this proves — and doesn't", font=f(11, bold=True, color=INK)); r += 1
    cell(r, 1, "Proves: the engine was graded on data it never saw, and its price→volume "
               "direction holds out-of-sample. Doesn't prove: a controlled causal effect — "
               "the forward moves were the brand's own, not a randomised test. A live price "
               "test (or Part B accumulating) closes that last gap.",
         font=f(9, italic=True, color=MUTED), align=al("left", wrap=True))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.row_dimensions[r].height = 48


def _build_leakage_sheet(ws, summary):
    """
    Leakage sheet — 'real vs borrowed vs stolen' uplift + the inelastic gate.
    All unit-based; no COGS/margin. Tells the brand which discounts are worth
    less than they look (high pull-forward / cannibalization) and which cells
    can't profit from discounting at all (inelastic).
    """
    lk = summary.get("leakage") or {}
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 34, "B": 13, "C": 14, "D": 16, "E": 12, "F": 14}.items():
        ws.column_dimensions[col].width = w

    def cell(r, c, val, font=None, align=None, fmt=None, border=None):
        x = ws.cell(row=r, column=c, value=val)
        x.font = font or f(10)
        if align: x.alignment = align
        if fmt: x.number_format = fmt
        if border: x.border = border
        return x

    cell(1, 1, "DISCOUNT OPTIMISATION  ·  LEAKAGE & DISCOUNT-WORTHINESS",
         font=f(9, color=MUTED), align=al("left"))
    cell(3, 1, "Real vs. Borrowed vs. Stolen", font=f(20, bold=True, color=INK))
    if not lk.get("available"):
        cell(5, 1, "Leakage analysis unavailable for this run.",
             font=f(10, italic=True, color=MUTED), align=al("left"))
        return
    cell(4, 1, "When a discount makes units jump, not all of it is new demand. "
               "BORROWED ≈ the dip below baseline in the weeks AFTER a promo (a sign of "
               "stock-up). STOLEN ≈ the dip in your own sibling packs DURING the promo. The "
               "rest is treated as REAL. These are estimated from observed dips — directional "
               "signals, not proven cause-and-effect. All unit-based, no margins.",
         font=f(10, color=BODY), align=al("left", wrap=True))
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=6)
    ws.row_dimensions[4].height = 42

    # headline counts
    cell(6, 1, f"Cells with promo history: {lk.get('n_with_promo',0)}/{lk.get('n_cells',0)}   ·   "
               f"typical real-demand share: {lk.get('median_true_incremental',1.0)*100:.0f}%   ·   "
               f"high-leakage cells (≥20% lost): {lk.get('n_high_leakage',0)}   ·   "
               f"inelastic cells (|ε|≤1, discount unlikely to pay): {lk.get('n_inelastic',0)}",
         font=f(10, bold=True, color=ACCENT), align=al("left"))
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=6)

    r = 8
    hdr = ["Product · City", "Borrowed (φ)", "Stolen (κ)", "Real demand",
           "|Elasticity|", "Worth discounting?"]
    for j, h in enumerate(hdr, 1):
        cell(r, j, h, font=f(10, bold=True, color=INK),
             align=al("right" if 1 < j < 6 else "left"),
             border=b(top=BOLD_RULE, bottom=RULE))
    r += 1
    # show the most-leaky / inelastic cells first; cap at 40 rows
    for c in lk.get("cells", [])[:40]:
        cell(r, 1, c["label"], align=al("left"))
        cell(r, 2, c["pull_forward"], fmt="0%", align=al("right"))
        cell(r, 3, c["cannibalization"], fmt="0%", align=al("right"))
        cell(r, 4, c["true_incremental_frac"], fmt="0%", align=al("right"))
        cell(r, 5, c["abs_elasticity"], fmt="0.00", align=al("right"))
        verdict = "Unlikely — inelastic, hold/raise" if c["is_inelastic"] else (
                  "weak — mostly leakage" if c["true_incremental_frac"] < 0.6 else "elastic — can work")
        vcolor = NEG if (c["is_inelastic"] or c["true_incremental_frac"] < 0.6) else POS
        cell(r, 6, verdict, font=f(9, bold=True, color=vcolor), align=al("left"))
        r += 1

    r += 1
    cell(r, 1, "How to use it: cells flagged 'inelastic' (|ε|≤1) are unlikely to profit from "
               "a discount — hold or raise price there. Cells with high Borrowed/Stolen % "
               "tend to give back most of the bump later or pull from your own packs — "
               "discount them shallower. These leakage haircuts are already applied to the "
               "(net) volume lift on the Price Drops list.",
         font=f(9, italic=True, color=MUTED), align=al("left", wrap=True))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 44


def _build_per_product_sheet(ws, summary, df, waste_main, reinvest_main):
    """
    Per-product sheet: each product gets a 1-line summary + a city-by-city
    table with the full week-by-week glide HORIZONTALLY.

    Layout per product:
       <Product title>
       Today: gross Rs.X, discount Rs.Y (Z%), N cities (a cut / b invest / c hold)
       ┌──────────┬─────────┬────────┬────────┬──────────┬───┬───┬───┬───┬───┐
       │ City     │ Current │ Target │ Action │ Save/mo  │W1 │W2 │W3 │...│W12│
       ├──────────┼─────────┼────────┼────────┼──────────┼───┼───┼───┼───┼───┤
       │ Bangalore│ 25.9%   │ 22.5%  │ CUT    │ Rs.8,247 │22.9│22.5│22.5│...│22.5│
       │ Mumbai   │ 24.7%   │ 22.6%  │ CUT    │ Rs.5,722 │22.6│22.6│22.6│...│22.6│
       │ ...                                                                  │
       └──────────────────────────────────────────────────────────────────────┘
    """
    import v4_config as cfg
    timeline = int(getattr(cfg, "TARGET_TIMELINE_WEEKS", 12))
    min_step = float(getattr(cfg, "MIN_DISCOUNT_CHANGE_PPT", 3))

    cut_set = set()
    if waste_main is not None and not waste_main.empty:
        cut_set = set(waste_main["cell_id"].dropna())
    inv_set = set()
    inv_rec = {}
    if reinvest_main is not None and not reinvest_main.empty:
        inv_set = set(reinvest_main["cell_id"].dropna())
        if "rec_discount_final" in reinvest_main.columns:
            inv_rec = (reinvest_main.dropna(subset=["rec_discount_final"])
                                    .set_index("cell_id")["rec_discount_final"]
                                    .to_dict())

    def cell(r, c, val, font=None, align=None, fmt=None, border=None):
        x = ws.cell(row=r, column=c, value=val)
        x.font = font or f(10)
        if align: x.alignment = align
        if fmt: x.number_format = fmt
        if border: x.border = border
        return x

    # ── Header ────────────────────────────────────────────────────────
    cell(1, 1, "DISCOUNT OPTIMISATION  ·  PER-PRODUCT  ·  CITY × WEEK MATRIX",
         font=f(9, color=MUTED), align=al("left"))
    cell(3, 1, "By product", font=f(20, bold=True, color=INK))
    cell(4, 1, f"Each product = one summary line, then one row per city showing the "
               f"planned discount % at every week of the {timeline}-week glide. "
               f"Action column: CUT = reduce discount (raise price), INVEST = increase "
               f"discount (drop price), HOLD = already at target.",
         font=f(10, color=BODY), align=al("left", wrap=True))
    n_cols = 8 + timeline  # 8 fixed (City, Conf, Cell R², Obs, CurRs, TgtRs, Action, Save) + N weekly
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=n_cols)
    ws.row_dimensions[4].height = 30

    # Sort products by product_id / grammage for deterministic order
    if "grammage" in df.columns:
        product_groups = list(df.groupby(["product_id", "grammage"], sort=True))
    else:
        product_groups = [(k, g) for k, g in df.groupby("product_id", sort=True)]

    cur_row = 7
    for key, prod_cells in product_groups:
        if isinstance(key, tuple):
            pid, grm = key
            grm = grm if pd.notna(grm) and str(grm).strip() else ""
        else:
            pid, grm = key, ""

        title = str(prod_cells["title"].iloc[0])[:60]
        display_title = f"{title}  ({grm})" if grm else title

        # ── Product summary line ──────────────────────────────────────
        gross   = float((prod_cells["mrp"] * prod_cells["current_units_day"] * 30).sum())
        spend   = float((prod_cells["mrp"] * prod_cells["current_discount_pct"] / 100
                          * prod_cells["current_units_day"] * 30).sum())
        wdisc   = (spend / gross * 100) if gross > 0 else 0
        n_cells = len(prod_cells)
        n_cut   = sum(1 for cid in prod_cells["cell_id"] if cid in cut_set and cid not in inv_set)
        n_inv   = sum(1 for cid in prod_cells["cell_id"] if cid in inv_set)
        n_hold  = n_cells - n_cut - n_inv

        # ── Product title row (merged) ────────────────────────────────
        cell(cur_row, 1, display_title,
             font=f(13, bold=True, color=INK), align=al("left"),
             border=b(bottom=BOLD_RULE))
        ws.merge_cells(start_row=cur_row, start_column=1,
                       end_row=cur_row, end_column=n_cols + 2)
        cur_row += 1

        # ── Mini-summary table in PROPER cells (label row + value row) ─
        # Two rows: labels on row N, values on row N+1. 8 columns of metrics.
        summary_labels = ["MRP",  "Today gross/mo", "Today discount/mo",
                          "Today disc %", "Cities", "Cut", "Invest", "Hold"]
        summary_values = [float(prod_cells["mrp"].iloc[0]),
                          gross, spend, wdisc,
                          n_cells, n_cut, n_inv, n_hold]
        for j, lbl in enumerate(summary_labels, 1):
            cell(cur_row, j, lbl,
                 font=f(9, color=MUTED), align=al("center"),
                 border=b(bottom=THIN))
        for j, val in enumerate(summary_values, 1):
            if j == 1:        # MRP
                fmt = '"Rs."#,##0'
            elif j in (2, 3): # gross / discount
                fmt = '"Rs."#,##0'
            elif j == 4:      # disc %
                fmt = '0.00"%"'
            else:             # counts
                fmt = "0"
            cell(cur_row + 1, j, val,
                 font=f(11, bold=True, color=INK), align=al("center"),
                 fmt=fmt, border=b(top=THIN))
        cur_row += 3  # 2 mini-summary rows + 1 blank

        # ── Accuracy note (cell, not a banner) ────────────────────────
        acc = summary.get("model_accuracy", {})
        if acc.get("available"):
            tier_color = {"Strong": POS, "Moderate": ACCENT,
                          "Weak": NEG, "Unreliable": NEG}.get(acc.get("tier", ""), MUTED)
            cell(cur_row, 1, "Model accuracy:",
                 font=f(9, color=MUTED), align=al("right"))
            cell(cur_row, 2, acc.get("tier", "—"),
                 font=f(10, bold=True, color=tier_color), align=al("left"))
            cell(cur_row, 3, "Price-engine R²:",
                 font=f(9, color=MUTED), align=al("right"))
            cell(cur_row, 4, acc.get("decision_r2_bin", acc.get("test_r2_agg", 0)),
                 font=f(10, bold=True), align=al("left"), fmt="0.00")
            cell(cur_row, 5, "MAPE:",
                 font=f(9, color=MUTED), align=al("right"))
            cell(cur_row, 6, acc.get("decision_mape_bin", acc.get("test_mape_agg", 0)),
                 font=f(10, bold=True), align=al("left"), fmt='0.0"%"')
            cur_row += 2  # 1 accuracy row + 1 blank

        # ── City × week table headers (selling-price view) ─────────────
        # Each row carries 3 confidence signals:
        #   Conf       — High/Medium/Low/Needs Experiment (Stage 5 tier)
        #   Cell R²    — model's IN-SAMPLE fit on THIS cell's own training data
        #                (a per-city fit signal, NOT out-of-sample validation)
        #   Obs        — number of training rows backing the estimate
        headers = ["City", "Conf", "Cell R²", "Obs",
                   "Current Rs.", "Target Rs.",
                   "Action", "Save Rs./mo"]
        headers += [f"W{w}" for w in range(1, timeline + 1)]
        for j, h in enumerate(headers, 1):
            cell(cur_row, j, h,
                 font=f(10, bold=True, color=INK),
                 align=al("center" if j > 1 else "left"),
                 border=b(top=BOLD_RULE, bottom=RULE))
        cur_row += 1

        # Compute per-cell glide and savings, then sort by savings desc
        city_rows = []
        for _, r in prod_cells.iterrows():
            cid = r.get("cell_id")
            city = str(r.get("city", ""))
            cur_d = float(r.get("current_discount_pct", 0))
            mrp = float(r.get("mrp", 0))
            cur_u = float(r.get("current_units_day", 0))

            # Action + target (discount %)
            if cid in inv_set:
                action = "INVEST"
                target_d = float(inv_rec.get(cid, cur_d + 3))
            elif cid in cut_set:
                action = "CUT"
                floor = float(r.get("historical_floor_disc", 0)) if pd.notna(r.get("historical_floor_disc")) else 0
                elbow = float(r.get("elbow_discount_pct", 0)) if pd.notna(r.get("elbow_discount_pct")) else 0
                target_d = max(elbow, floor)
            else:
                action = "HOLD"
                target_d = cur_d

            # Per-cycle step (same rule as Stage 7)
            gap = abs(cur_d - target_d)
            if gap < 0.1:
                step = 0.0
            elif gap <= min_step:
                step = gap
            else:
                step = max(min_step, gap / float(timeline))
            direction = -1 if (cur_d - target_d) > 0 else (1 if (cur_d - target_d) < 0 else 0)

            # Per-week DISCOUNT % then convert to selling PRICE
            weekly_prices = []
            for w in range(1, timeline + 1):
                if direction == 0:
                    d = cur_d
                else:
                    d = cur_d + direction * step * w
                    d = max(d, target_d) if direction < 0 else min(d, target_d)
                price_w = mrp * (1 - d / 100)
                weekly_prices.append(round(price_w, 1))

            # Convert current/target discount to selling price
            cur_price    = mrp * (1 - cur_d / 100)
            target_price = mrp * (1 - target_d / 100)

            # Monthly savings (Rs.) — same calc, in discount-spend terms
            save_per_mo = (cur_d - target_d) / 100 * mrp * cur_u * 30

            # Per-cell confidence (from Stage 5) + observation count
            conf = str(r.get("confidence", "—"))
            n_obs = int(r.get("n_observations", 0)) if pd.notna(r.get("n_observations")) else 0
            cell_r2 = float(r.get("cell_train_r2", 0)) if pd.notna(r.get("cell_train_r2")) else 0.0

            city_rows.append({
                "city": city,
                "conf": conf, "cell_r2": cell_r2, "n_obs": n_obs,
                "cur_price": cur_price, "target_price": target_price,
                "action": action, "save": save_per_mo,
                "weekly_prices": weekly_prices,
                "sort_key": abs(save_per_mo),
            })

        city_rows.sort(key=lambda x: -x["sort_key"])

        # Render rows — all monetary values
        # Column layout: City | Conf | Cell R² | Obs | Cur | Tgt | Action | Save | W1..Wn
        total_cols_per_row = 8 + timeline
        for r in city_rows:
            action_color = NEG if r["action"] == "CUT" else (POS if r["action"] == "INVEST" else MUTED)
            conf_color = {
                "High":              POS,
                "Medium":            ACCENT,
                "Low":               NEG,
                "Needs Experiment":  NEG,
            }.get(r["conf"], MUTED)
            # Cell R² color: green ≥0.7, slate ≥0.4, red ≥0.1, grey below
            r2v = r["cell_r2"]
            if r2v >= 0.7:    r2_color = POS
            elif r2v >= 0.4:  r2_color = ACCENT
            elif r2v >= 0.1:  r2_color = NEG
            else:             r2_color = MUTED

            cell(cur_row, 1, r["city"], font=f(10), align=al("left"))
            cell(cur_row, 2, r["conf"],
                 font=f(10, bold=True, color=conf_color), align=al("center"))
            cell(cur_row, 3, r2v,
                 font=f(10, bold=True, color=r2_color), align=al("center"),
                 fmt="0.00")
            cell(cur_row, 4, r["n_obs"], font=f(10), align=al("center"), fmt="#,##0")
            cell(cur_row, 5, r["cur_price"], font=f(10), align=al("right"),
                 fmt='"Rs."#,##0.0')
            cell(cur_row, 6, r["target_price"], font=f(10, bold=True), align=al("right"),
                 fmt='"Rs."#,##0.0')
            cell(cur_row, 7, r["action"],
                 font=f(10, bold=True, color=action_color), align=al("center"))
            sav = r["save"]
            sav_color = POS if sav > 0 else (NEG if sav < 0 else MUTED)
            cell(cur_row, 8, sav, font=f(10, color=sav_color), align=al("right"),
                 fmt='"Rs. "#,##0;"Rs. -"#,##0;"-"')

            # Weekly price cells start at column 9
            for wi, wprice in enumerate(r["weekly_prices"], 1):
                col = 8 + wi
                reached = abs(wprice - r["target_price"]) < 0.05
                if r["action"] == "HOLD":
                    wfont = f(9, color=MUTED)
                elif reached:
                    wfont = f(9, bold=True, color=NEG if r["action"] == "CUT" else POS)
                else:
                    wfont = f(9, color=BODY)
                cell(cur_row, col, wprice, font=wfont, align=al("center"),
                     fmt='#,##0.0')

            for c in range(1, total_cols_per_row + 1):
                ws.cell(row=cur_row, column=c).border = b(bottom=THIN)
            cur_row += 1

        for c in range(1, total_cols_per_row + 1):
            ws.cell(row=cur_row - 1, column=c).border = b(bottom=BOLD_RULE)
        cur_row += 2

    # Column widths (City + Conf + Cell R² + Obs + 2 prices + Action + Save + N weekly)
    widths = [22, 11, 9, 7, 13, 13, 10, 14] + [9] * timeline
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze the leftmost 8 columns (through Save Rs./mo) so city + conf +
    # cell R² + obs + prices + action stay visible while scrolling right.
    ws.freeze_panes = "I8"


def _build_detail_sheet(ws, df, sheet_type):
    """
    Per-cell detail tables. These are values (not formulas) — they're a
    snapshot of which cells got which recommendation, sorted for action.
    """
    def cell(r, c, val, font=None, align=None, fmt=None, border=None):
        x = ws.cell(row=r, column=c, value=val)
        x.font = font or f(10)
        if align: x.alignment = align
        if fmt:   x.number_format = fmt
        if border:x.border = border
        return x

    if sheet_type == "cut":
        title = "Where to raise prices this week"
        eyebrow = "DISCOUNT OPTIMISATION  ·  PRICE LIFTS"
        desc = ("Cells sorted by Rs. wasted per month. "
                "Now is the current selling price. "
                "This Week is what to set on Blinkit this Monday — capped at a 3 ppt move "
                "per cycle. Wasted/mo is the full multi-cycle savings opportunity if you "
                "walked the price all the way back to MRP.")
        cols = [
            ("title",            "Product",         40),
            ("city",             "City",            20),
            ("mrp",              "MRP",             10),
            ("current_price",    "Now",             12),
            ("this_week_price",  "This Week",       14),
            ("wasted_inr_per_month", "Wasted/mo",   16),
            ("confidence",       "Conf",            12),
        ]
        money_cols = {"mrp", "current_price", "this_week_price", "wasted_inr_per_month"}
        pct_cols   = set()
    elif sheet_type == "invest":
        title = "Where to drop prices to grow volume"
        eyebrow = "DISCOUNT OPTIMISATION  ·  STRATEGIC INVESTMENTS"
        desc = ("Cells where dropping the price by 3 ppt is projected to add enough volume "
                "to be worth the extra discount spend. These are funded by the savings from "
                "the price lifts above.")
        cols = [
            ("title",                       "Product",       40),
            ("city",                        "City",          18),
            ("mrp",                         "MRP",           10),
            ("current_price",               "Now",           11),
            ("new_price",                   "New",           11),
            ("volume_lift_pct",             "Vol Δ",         11),
            ("extra_volume_units_per_month","+Units/mo",     14),
            ("budget_needed_inr_per_month", "Budget/mo",     14),
            ("confidence",                  "Conf",          12),
        ]
        money_cols = {"mrp", "current_price", "new_price",
                      "budget_needed_inr_per_month", "extra_volume_units_per_month"}
        pct_cols   = {"volume_lift_pct"}
    else:  # needs_test
        title = "Cells needing a price test"
        eyebrow = "DISCOUNT OPTIMISATION  ·  PILOT REQUIRED"
        desc = ("These cells don't have enough clean data to act on. The model isn't "
                "confident enough — usually because of too few observations, too little "
                "price variation, or a launch ramp confounding the elasticity signal. "
                "Run a small A/B test in one city before changing anything.")
        cols = [
            ("title",                "Product",     60),
            ("city",                 "City",        24),
            ("current_discount_pct", "Now %",       12),
            ("elbow_discount_pct",   "Model %",     12),
            ("confidence",           "Status",      20),
        ]
        money_cols = set()
        pct_cols   = {"current_discount_pct", "elbow_discount_pct"}

    cell(1, 1, eyebrow, font=f(9, color=MUTED), align=al("left"))
    cell(3, 1, title, font=f(20, bold=True, color=INK))
    cell(4, 1, desc, font=f(10, color=BODY), align=al("left", wrap=True))
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=len(cols))
    ws.row_dimensions[4].height = 32

    # Confidence legend
    cell(6, 1,
         "Confidence: High = 200+ days of clean history with 10+ distinct discount "
         "levels and ≥3 ppt of price variation; the model trusts this cell. "
         "Medium = 100+ days and 5+ discount levels — actionable but worth a review. "
         "Low = thin data; shown separately as 'Needs Price Test'. "
         "Cells with a data-quality concern flagged (boundary-hit elasticity or rapid "
         "demand growth) are automatically downgraded one tier.",
         font=f(8, italic=True, color=MUTED), align=al("left", wrap=True))
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=len(cols))
    ws.row_dimensions[6].height = 40

    # Header row
    H = 8
    for j, (_, label, _) in enumerate(cols, 1):
        cell(H, j, label,
             font=f(10, bold=True, color=INK),
             align=al("right" if j > 2 else "left"),
             border=b(top=BOLD_RULE, bottom=RULE))

    # Data rows
    if df is None or df.empty:
        cell(H + 1, 1, "(no cells)", font=f(10, italic=True, color=MUTED), align=al("left"))
    else:
        r = H + 1
        for _, row in df.head(80).iterrows():
            for j, (col, _, _) in enumerate(cols, 1):
                v = row.get(col, "")
                if pd.isna(v):
                    cell(r, j, "", border=b(bottom=THIN))
                    continue
                if col == "title":
                    cell(r, j, str(v)[:60], font=f(10),
                         align=al("left", wrap=True), border=b(bottom=THIN))
                elif col in money_cols:
                    cell(r, j, float(v), font=f(10), align=al("right"),
                         fmt="#,##0", border=b(bottom=THIN))
                elif col in pct_cols:
                    cell(r, j, float(v), font=f(10), align=al("right"),
                         fmt='+0.0"%";-0.0"%"' if col == "volume_lift_pct" else '0.0"%"',
                         border=b(bottom=THIN))
                elif isinstance(v, (int, float)):
                    cell(r, j, float(v), font=f(10), align=al("right"),
                         fmt="0.0", border=b(bottom=THIN))
                else:
                    cell(r, j, str(v), font=f(10),
                         align=al("center" if col == "confidence" else "left"),
                         border=b(bottom=THIN))
            r += 1

    # Column widths
    for j, (_, _, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = f"A{H + 1}"


# ────────────────────────────────────────────────────────────────────
# Helpers
# ────────────────────────────────────────────────────────────────────
def _get_target_disc(summary):
    """Fetch target_weighted_discount_pct from config, defaulting to 9.0."""
    try:
        import v4_config as cfg
        return float(cfg.TARGET_WEIGHTED_DISCOUNT_PCT)
    except Exception:
        return 9.0


def _filter_actual_cuts(df, waste_main, reinvest_main):
    """Count cells that are actually cut (in waste but NOT in reinvest)."""
    if waste_main is None or waste_main.empty:
        return pd.DataFrame()
    if reinvest_main is None or reinvest_main.empty:
        return waste_main
    return waste_main[~waste_main["cell_id"].isin(reinvest_main["cell_id"])]
