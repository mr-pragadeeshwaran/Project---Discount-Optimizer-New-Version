"""
settings_loader.py — brand settings from a FILE, not from code.

WHY
---
Every knob a brand/engagement owns (savings target, budget cap, hero SKUs,
cost structure, festival calendar) used to be edited inside v4_config.py.
That put a non-engineer one typo away from breaking an import, and made
"what settings produced this plan?" a git-archaeology question.

HOW IT WORKS
------------
  1. v4_config.py still defines every DEFAULT, exactly as before.
  2. If an override file exists, the values in it win.
  3. Nothing else changes: every consumer keeps doing `import v4_config as cfg`
     and reading `cfg.WHATEVER`. Overrides are applied at import time.

Override file — either format, both optional:
     config/settings.xlsx   sheets: Settings | Festivals | Platform Events
     config/settings.csv    + config/festivals.csv + config/platform_events.csv
xlsx wins if both are present (a warning is printed).

File rules:
  - Only the `key` and `value` columns are read; unit/description are there
    to make the template self-documenting.
  - A BLANK value means "use the code default" — it is not an empty string.
  - To set an explicitly empty list, write `none` or `[]`.
  - An unknown key is an ERROR, not a silent no-op (catches typos like
    SAVINGS_TARGET_MONTHY_INR).
  - A Festivals/Platform Events sheet with rows REPLACES the code calendar
    outright (the template ships pre-filled with the current calendar, so
    what you see in the file is what the system uses).

The template is GENERATED from the REGISTRY below, so it can never drift
from what the code actually supports.

Fail-loud by design: a bad settings file raises SettingsError on import of
v4_config. A wrong number that runs silently is far more expensive than a
crash that names the offending cell.
"""
import os
import datetime

HERE = os.path.dirname(os.path.abspath(__file__))
CONFIG_DIR = os.path.join(HERE, "config")

XLSX_PATH = os.path.join(CONFIG_DIR, "settings.xlsx")
CSV_PATH = os.path.join(CONFIG_DIR, "settings.csv")
FESTIVALS_CSV = os.path.join(CONFIG_DIR, "festivals.csv")
EVENTS_CSV = os.path.join(CONFIG_DIR, "platform_events.csv")

SHEET_SETTINGS = "Settings"
SHEET_FESTIVALS = "Festivals"
SHEET_EVENTS = "Platform Events"


class SettingsError(Exception):
    """Raised when the settings file is present but cannot be trusted."""


# ── Types ──────────────────────────────────────────────────────────────────
# Each parser takes the raw cell text and returns a typed value, or raises
# ValueError with a message written for the person editing the spreadsheet.

def _p_text(v):
    return str(v).strip()


def _p_int(v):
    s = str(v).strip().replace(",", "").replace("_", "")
    f = float(s)                      # tolerate "500000.0" from Excel
    if abs(f - round(f)) > 1e-9:
        raise ValueError(f"must be a whole number, got {v!r}")
    return int(round(f))


def _p_number(v):
    return float(str(v).strip().replace(",", "").replace("_", ""))


def _p_fraction(v):
    """0-1 fraction. Rejects the classic 12-vs-0.12 percent mix-up."""
    f = _p_number(v)
    if not 0.0 <= f <= 1.0:
        raise ValueError(
            f"must be a FRACTION between 0 and 1 (0.12 = 12%), got {v!r}"
            + (f" — did you mean {f/100:g}?" if 1 < f <= 100 else "")
        )
    return f


def _p_percent(v):
    f = _p_number(v)
    if not 0.0 <= f <= 100.0:
        raise ValueError(f"must be a PERCENT between 0 and 100 (12 = 12%), got {v!r}")
    return f


def _p_yes_no(v):
    s = str(v).strip().lower()
    if s in ("yes", "y", "true", "1"):
        return True
    if s in ("no", "n", "false", "0"):
        return False
    raise ValueError(f"must be yes or no, got {v!r}")


def _split_list(v):
    s = str(v).strip()
    if s.lower() in ("none", "[]", "-"):
        return []
    # accept pipe- or comma-separated; pipe wins so values may contain commas
    parts = s.split("|") if "|" in s else s.split(",")
    return [p.strip() for p in parts if p.strip()]


def _p_list_text(v):
    return _split_list(v)


def _p_list_id(v):
    """SKU/product ids. Kept as strings; Excel loves turning 532393 into
    532393.0, and downstream _clean_pid matching is string-based."""
    out = []
    for p in _split_list(v):
        try:
            f = float(p)
            p = str(int(round(f))) if abs(f - round(f)) < 1e-9 else p
        except ValueError:
            pass
        out.append(p)
    return out


PARSERS = {
    "text": _p_text, "integer": _p_int, "number": _p_number,
    "fraction": _p_fraction, "percent": _p_percent, "yes_no": _p_yes_no,
    "list_text": _p_list_text, "list_id": _p_list_id,
}

UNIT_HELP = {
    "text": "text", "integer": "whole number", "number": "number",
    "fraction": "fraction 0-1 (0.12 = 12%)", "percent": "percent 0-100 (12 = 12%)",
    "yes_no": "yes / no", "list_text": "list, separate with |",
    "list_id": "list of ids, separate with |",
}


# ── The registry: every knob a brand owner may set from the file ───────────
# (key, type, section, description). Defaults are NOT here — they live in
# v4_config.py and are read from it, so there is exactly one source of truth.
REGISTRY = [
    # ── Business targets ──
    ("SAVINGS_TARGET_MONTHLY_INR", "integer", "Business targets",
     "Monthly savings ambition in Rs. Gate C6 reports MEETS/BELOW against this. "
     "A verdict only — a smaller plan still executes."),
    ("TARGET_DISCOUNT_PCT", "percent", "Business targets",
     "Headline discount % the brand is steering toward (dashboard KPI)."),
    ("TARGET_WEIGHTED_DISCOUNT_PCT", "percent", "Business targets",
     "Revenue-weighted portfolio discount % the flywheel glides toward."),
    ("TARGET_QUARTER", "text", "Business targets",
     "Quarter label the discount target is set for, e.g. Q4 2026."),
    ("DEFAULT_BUDGET_PCT_CAP", "fraction", "Business targets",
     "Weekly discount-spend cap as a fraction of gross sales. 0.12 = 12%. "
     "Blocks reinvestment when breached; never forces extra cuts."),
    ("TARGET_TIMELINE_WEEKS", "integer", "Business targets",
     "Every cell's full discount gap must close within this many weekly cycles."),
    ("MIN_DISCOUNT_CHANGE_PPT", "number", "Business targets",
     "Smallest weekly discount step in percentage points. Gaps smaller than "
     "this close in one cycle."),

    # ── Brand identity ──
    ("BRAND_NAME", "text", "Brand identity",
     "The client's own brand as it appears on the platform."),
    ("OWN_BRAND_PATTERNS", "list_text", "Brand identity",
     "Every spelling of the own brand in the data. Blank = derive from "
     "BRAND_NAME. Ingestion fails loud if these match no rows, or match a "
     "competitor too."),
    ("PLATFORM_NAME", "text", "Brand identity",
     "The quick-commerce platform, e.g. Blinkit."),
    ("STRATEGIC_SKUS", "list_id", "Brand identity",
     "Hero PRODUCT_IDs that must NEVER be auto-cut, whatever the model says. "
     "Separate with |. Write none for no hero protection."),
    ("STRICT_OWN_BRAND_MATCH", "yes_no", "Brand identity",
     "yes = fail loudly if the brand patterns also match a competitor brand "
     "(e.g. 'Sun' catching 'Sunfeast'). Keep yes unless you truly own several "
     "distinct brand strings."),

    # ── Cost structure ──
    ("DEFAULT_COGS_PCT", "fraction", "Cost structure",
     "Cost of goods as a fraction of MRP. 0.50 = 50%. PROXY until finance "
     "supplies real per-SKU costs — every profit figure inherits this."),
    ("DEFAULT_COMMISSION_PCT", "fraction", "Cost structure",
     "Platform commission as a fraction of selling price. 0.15 = 15%."),
    ("DEFAULT_FULFILLMENT_FEE", "number", "Cost structure",
     "Fulfilment cost in Rs per unit."),

    # ── Model & guardrails ──
    ("TRAIN_LOOKBACK_DAYS", "integer", "Model and guardrails",
     "Train only on the last N days of regular trading. The single biggest "
     "accuracy lever — shorter avoids stale price regimes and launch ramps."),
    ("OUTLIER_Z_THRESHOLD", "number", "Model and guardrails",
     "Per-cell z-score above which a day is treated as an outlier and excluded "
     "from training. 2.0 is the tuned production value; 3.0 keeps more days."),
    ("MARGINAL_ROI_THRESHOLD", "number", "Model and guardrails",
     "The elbow: deepest discount whose marginal ROI still clears this. 1.0 = "
     "the last rupee of discount must at least pay for itself."),
    ("HISTORICAL_FLOOR_PERCENTILE", "number", "Model and guardrails",
     "Percentile of a cell's own recent discounts treated as its proven-safe "
     "floor. 25 = 'we ran this low on a quarter of days and survived'."),
    ("INELASTIC_ELASTICITY_THRESHOLD", "number", "Model and guardrails",
     "Cells with |elasticity| at or below this cannot pay for a discount; they "
     "are held/raised, never reinvested. 1.0 is the theorem boundary."),
    ("REINVEST_MIN_VOL_LIFT_PCT", "percent", "Model and guardrails",
     "A reinvest candidate must lift volume by at least this %, net of leakage."),
    ("REINVEST_MAX_MARGIN_SAC_PCT", "percent", "Model and guardrails",
     "A reinvest candidate may sacrifice at most this % of current contribution."),
    ("REINVEST_MIN_ELASTICITY", "number", "Model and guardrails",
     "|elasticity| a cell needs before deeper discount is even considered."),
    ("VOLUME_DROP_TOLERANCE_PCT", "percent", "Model and guardrails",
     "Kill-switch: a cell strikes when actual volume misses prediction by more "
     "than this %. Two strikes revert the cut."),
    ("DRIFT_ALERT_THRESHOLD", "fraction", "Model and guardrails",
     "Prediction error fraction that counts as model drift. 0.15 = 15%."),
]

REGISTRY_BY_KEY = {k: (t, s, d) for k, t, s, d in REGISTRY}
SECTIONS = list(dict.fromkeys(s for _, _, s, _ in REGISTRY))


# ── Reading ────────────────────────────────────────────────────────────────
# Everything is read as plain rows-of-cells first (csv or xlsx), then parsed by
# ONE tabular reader that SCANS for the header row rather than assuming row 1.
# The templates carry a human note above the header and people add notes of
# their own: a file we hand out must be a file we can read back.

def _csv_rows(path):
    import csv
    with open(path, newline="", encoding="utf-8-sig") as fh:
        return [[("" if c is None else c) for c in row] for row in csv.reader(fh)]


def _sheet_rows(wb, name):
    if name not in wb.sheetnames:
        return None
    return [["" if c is None else c for c in row]
            for row in wb[name].iter_rows(values_only=True)]


def _tabular(rows, needed, where):
    """[{col: cell}] for every data row beneath the first row that carries all
    of `needed`. Extra columns (unit/description) are ignored, notes above the
    header are skipped, blank spacer rows are dropped."""
    if not rows:
        return []
    for i, row in enumerate(rows):
        header = [str(c).strip().lower() for c in row]
        if all(n in header for n in needed):
            idx = {n: header.index(n) for n in needed}
            out = []
            for r in rows[i + 1:]:
                if all(str(c).strip() == "" for c in r):
                    continue
                out.append({n: (r[j] if j < len(r) else "") for n, j in idx.items()})
            return out
    raise SettingsError(
        f"{where}: no header row found with "
        + " and ".join(f"'{n}'" for n in needed)
        + f" columns (first row seen: {rows[0]}). Download a fresh template from "
          f"the dashboard's Inputs & Settings page."
    )


def _tabular_pairs(rows, col_a, col_b, where):
    return [(r[col_a], r[col_b]) for r in _tabular(rows, (col_a, col_b), where)]


def _norm_date(v, where):
    """Excel hands back datetimes; text files hand back strings."""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return ""
    try:
        return datetime.date.fromisoformat(s[:10]).isoformat()
    except ValueError:
        raise SettingsError(f"{where}: '{s}' is not a date in YYYY-MM-DD form.")


def _read_files():
    """-> (settings_pairs, festival_pairs|None, event_triples|None, source_label)
    None for a calendar means 'not supplied — keep the code default'."""
    have_xlsx, have_csv = os.path.exists(XLSX_PATH), os.path.exists(CSV_PATH)
    if not have_xlsx and not have_csv:
        return [], None, None, None

    if have_xlsx and have_csv:
        print("[settings] WARNING: both config/settings.xlsx and config/settings.csv "
              "exist — using the .xlsx and IGNORING the .csv.")

    def _events_from(rows, where):
        if rows is None:
            return None
        return [(r["start"], r["end"], r["event"])
                for r in _tabular(rows, ("start", "end", "event"), where)]

    if have_xlsx:
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise SettingsError(
                "config/settings.xlsx found but openpyxl is not installed "
                "(pip install -r requirements.txt), or use config/settings.csv instead."
            )
        wb = load_workbook(XLSX_PATH, data_only=True)
        if SHEET_SETTINGS not in wb.sheetnames:
            raise SettingsError(
                f"config/settings.xlsx has no '{SHEET_SETTINGS}' sheet "
                f"(found: {wb.sheetnames}). Download a fresh template."
            )
        settings = _tabular_pairs(_sheet_rows(wb, SHEET_SETTINGS), "key", "value",
                                  f"settings.xlsx / {SHEET_SETTINGS}")
        fest_rows = _sheet_rows(wb, SHEET_FESTIVALS)
        fests = (_tabular_pairs(fest_rows, "date", "event",
                                f"settings.xlsx / {SHEET_FESTIVALS}")
                 if fest_rows is not None else None)
        events = _events_from(_sheet_rows(wb, SHEET_EVENTS),
                              f"settings.xlsx / {SHEET_EVENTS}")
        return settings, fests, events, "config/settings.xlsx"

    settings = _tabular_pairs(_csv_rows(CSV_PATH), "key", "value",
                              "config/settings.csv")
    fests = (_tabular_pairs(_csv_rows(FESTIVALS_CSV), "date", "event",
                            "config/festivals.csv")
             if os.path.exists(FESTIVALS_CSV) else None)
    events = (_events_from(_csv_rows(EVENTS_CSV), "config/platform_events.csv")
              if os.path.exists(EVENTS_CSV) else None)
    return settings, fests, events, "config/settings.csv"


def _parse_settings(pairs, source):
    """-> {key: typed value} for non-blank rows. Raises on typo/type errors,
    collecting every problem so one edit round fixes them all."""
    out, errors = {}, []
    for raw_key, raw_val in pairs:
        key = str(raw_key or "").strip()
        if not key or key.startswith("#"):
            continue
        if key not in REGISTRY_BY_KEY:
            near = [k for k in REGISTRY_BY_KEY if k.split("_")[0] == key.split("_")[0]]
            errors.append(f"  '{key}' is not a setting this system has"
                          + (f" — did you mean {near[0]}?" if near else ""))
            continue
        if raw_val is None or str(raw_val).strip() == "":
            continue                     # blank = keep the code default
        typ = REGISTRY_BY_KEY[key][0]
        try:
            out[key] = PARSERS[typ](raw_val)
        except ValueError as e:
            errors.append(f"  {key}: {e}")
    if errors:
        raise SettingsError(
            f"{source} has {len(errors)} problem(s):\n" + "\n".join(errors)
            + "\n\nFix the file (or download a fresh template from the dashboard's "
              "Inputs & Settings page) and run again. Nothing was changed."
        )
    return out


def _parse_calendar(pairs, source):
    if pairs is None:
        return None
    out = {}
    for d, name in pairs:
        ds = _norm_date(d, source)
        if not ds:
            continue
        out[ds] = str(name).strip() or "Event"
    return out or None          # an empty sheet means "not supplied"


def _parse_events(triples, source):
    if triples is None:
        return None
    out = {}
    for s, e, name in triples:
        ss, es = _norm_date(s, source), _norm_date(e, source)
        if not ss or not es:
            continue
        if es < ss:
            raise SettingsError(f"{source}: event '{name}' ends ({es}) before it "
                                f"starts ({ss}).")
        out[(ss, es)] = str(name).strip() or "Event"
    return out or None


# ── Public API ─────────────────────────────────────────────────────────────
STATE = {"source": None, "overrides": {}, "calendar_source": None, "error": None}


def apply_to(ns):
    """Apply file overrides onto a namespace dict (v4_config's globals()).

    Called once, from the bottom of v4_config.py. Records what happened in
    STATE so the dashboard can show which values came from the file.
    """
    settings, fests, events, source = _read_files()
    if source is None:
        STATE.update(source=None, overrides={}, calendar_source=None, error=None)
        return {}

    values = _parse_settings(settings, source)
    for k, v in values.items():
        ns[k] = v

    cal_src = []
    fest_map = _parse_calendar(fests, source)
    if fest_map:
        ns["FESTIVAL_DATES"] = fest_map
        cal_src.append(f"{len(fest_map)} festival dates")
    ev_map = _parse_events(events, source)
    if ev_map:
        ns["PLATFORM_EVENT_WINDOWS"] = ev_map
        cal_src.append(f"{len(ev_map)} platform events")

    STATE.update(source=source, overrides=values,
                 calendar_source=", ".join(cal_src) or None, error=None)
    if values or cal_src:
        bits = f"{len(values)} setting(s)" + (f" + {', '.join(cal_src)}" if cal_src else "")
        print(f"[settings] {source}: {bits} override the code defaults.")
    return values


def describe():
    """[{key, section, unit, description, value, source}] — effective config as
    the dashboard shows it. Reads live values from v4_config."""
    import v4_config as cfg
    rows = []
    for key, typ, section, desc in REGISTRY:
        val = getattr(cfg, key, None)
        rows.append({
            "key": key, "section": section, "type": typ,
            "unit": UNIT_HELP.get(typ, typ), "description": desc,
            "value": val if not isinstance(val, (list, tuple)) else list(val),
            "source": "file" if key in STATE.get("overrides", {}) else "default",
        })
    return rows


def status():
    """What the dashboard's settings card shows."""
    import v4_config as cfg
    return {
        "source": STATE.get("source"),
        "n_overrides": len(STATE.get("overrides", {})),
        "calendar_source": STATE.get("calendar_source"),
        "error": STATE.get("error"),
        "xlsx_exists": os.path.exists(XLSX_PATH),
        "csv_exists": os.path.exists(CSV_PATH),
        "festival_dates": len(getattr(cfg, "FESTIVAL_DATES", {}) or {}),
        "festival_last": max((getattr(cfg, "FESTIVAL_DATES", {}) or {"": ""}).keys() or [""]),
    }


# ── Template generation (from the registry — cannot drift) ─────────────────
def _template_rows():
    import v4_config as cfg
    rows = []
    for key, typ, section, desc in REGISTRY:
        val = getattr(cfg, key, "")
        if isinstance(val, (list, tuple)):
            val = " | ".join(str(x) for x in val) if val else "none"
        elif isinstance(val, bool):
            val = "yes" if val else "no"
        rows.append([section, key, val, UNIT_HELP.get(typ, typ), desc])
    return rows


def _calendar_rows():
    import v4_config as cfg
    fest = [[d, n] for d, n in sorted((getattr(cfg, "FESTIVAL_DATES", {}) or {}).items())]
    ev = [[s, e, n] for (s, e), n in
          sorted((getattr(cfg, "PLATFORM_EVENT_WINDOWS", {}) or {}).items())]
    return fest, ev


HEADER_NOTE = (
    "Edit the value column only. Blank value = keep the built-in default. "
    "Write 'none' to empty a list. Unknown keys are rejected."
)


def template_csv():
    """The Settings template as CSV text."""
    import csv, io
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\n")
    w.writerow(["# " + HEADER_NOTE])
    w.writerow(["section", "key", "value", "unit", "description"])
    for r in _template_rows():
        w.writerow(r)
    return buf.getvalue()


def template_festivals_csv():
    import csv, io
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\n")
    w.writerow(["date", "event"])
    fest, _ = _calendar_rows()
    for r in fest:
        w.writerow(r)
    return buf.getvalue()


def template_events_csv():
    import csv, io
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\n")
    w.writerow(["start", "end", "event"])
    _, ev = _calendar_rows()
    for r in ev:
        w.writerow(r)
    return buf.getvalue()


def template_xlsx_bytes():
    """One workbook: Settings + Festivals + Platform Events, styled enough to
    be obvious about which column to edit."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="1E293B")
    head_font = Font(color="FFFFFF", bold=True)
    edit_fill = PatternFill("solid", fgColor="FEF9C3")
    sec_font = Font(bold=True)

    ws = wb.active
    ws.title = SHEET_SETTINGS
    ws.append(["section", "key", "value", "unit", "description"])
    for c in ws[1]:
        c.fill, c.font = head_fill, head_font
    last_section = None
    for section, key, val, unit, desc in _template_rows():
        ws.append([section, key, val, unit, desc])
        r = ws.max_row
        ws.cell(r, 3).fill = edit_fill          # the value column is the editable one
        if section != last_section:
            ws.cell(r, 1).font = sec_font
            last_section = section
    for col, width in zip("ABCDE", (22, 34, 30, 26, 86)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.insert_rows(1)
    ws["A1"] = HEADER_NOTE
    ws["A1"].font = Font(italic=True, color="475569")
    ws.freeze_panes = "A3"

    fest, ev = _calendar_rows()
    wf = wb.create_sheet(SHEET_FESTIVALS)
    wf.append(["date", "event"])
    for c in wf[1]:
        c.fill, c.font = head_fill, head_font
    for r in fest:
        wf.append(r)
    wf.column_dimensions["A"].width = 14
    wf.column_dimensions["B"].width = 34
    wf.freeze_panes = "A2"

    we = wb.create_sheet(SHEET_EVENTS)
    we.append(["start", "end", "event"])
    for c in we[1]:
        c.fill, c.font = head_fill, head_font
    for r in ev:
        we.append(r)
    for col, width in zip("ABC", (14, 14, 34)):
        we.column_dimensions[col].width = width
    we.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def validate_bytes(filename, data):
    """Dry-run a candidate settings file WITHOUT installing it.

    Returns (ok, message). Used by the dashboard's upload flow so a bad file
    is rejected before it can break the next run.
    """
    import tempfile, shutil
    ext = os.path.splitext(filename)[1].lower()
    if ext not in (".csv", ".xlsx"):
        return False, f"'{filename}' must be a .csv or .xlsx file."
    tmpdir = tempfile.mkdtemp(prefix="statiq_settings_")
    try:
        # Point the module at a throwaway dir, parse there, then restore.
        global XLSX_PATH, CSV_PATH, FESTIVALS_CSV, EVENTS_CSV
        keep = (XLSX_PATH, CSV_PATH, FESTIVALS_CSV, EVENTS_CSV)
        probe = os.path.join(tmpdir, "settings" + ext)
        with open(probe, "wb") as fh:
            fh.write(data)
        XLSX_PATH = probe if ext == ".xlsx" else os.path.join(tmpdir, "none.xlsx")
        CSV_PATH = probe if ext == ".csv" else os.path.join(tmpdir, "none.csv")
        FESTIVALS_CSV = os.path.join(tmpdir, "festivals.csv")
        EVENTS_CSV = os.path.join(tmpdir, "platform_events.csv")
        try:
            settings, fests, events, source = _read_files()
            values = _parse_settings(settings, filename)
            _parse_calendar(fests, filename)
            _parse_events(events, filename)
        finally:
            XLSX_PATH, CSV_PATH, FESTIVALS_CSV, EVENTS_CSV = keep
    except SettingsError as e:
        return False, str(e)
    except Exception as e:
        return False, f"Could not read '{filename}': {e}"
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)
    return True, f"{len(values)} setting(s) will override the defaults."


def install_bytes(filename, data):
    """Validate, then write into config/ as the live settings file.

    Only ever writes to the fixed config/settings.{csv,xlsx} path — the
    uploaded filename is never used as a destination.
    """
    ok, msg = validate_bytes(filename, data)
    if not ok:
        return False, msg
    ext = os.path.splitext(filename)[1].lower()
    os.makedirs(CONFIG_DIR, exist_ok=True)
    dest = XLSX_PATH if ext == ".xlsx" else CSV_PATH
    other = CSV_PATH if ext == ".xlsx" else XLSX_PATH
    with open(dest, "wb") as fh:
        fh.write(data)
    note = ""
    if os.path.exists(other):
        os.remove(other)
        note = f" (removed the old {os.path.basename(other)} so there is one source of truth)"
    return True, (f"Saved to config/{os.path.basename(dest)}{note}. {msg} "
                  f"They take effect on the next run.")
